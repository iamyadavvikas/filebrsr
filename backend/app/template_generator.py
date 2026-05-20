"""
BRSR Template Generator — Generates SEBI-format BRSR Excel file from extracted data.

Produces an Excel workbook with:
- Section A: General Disclosures
- Section B: Management & Process (9 principles policy table)
- Section C: Principle-wise Performance (P1 through P9)

Uses openpyxl for .xlsx generation (lighter than pandas for this use case).
"""

import io
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════
# TEMPLATE STRUCTURE
# ═══════════════════════════════════════════════════════════════════════

SECTION_A_TEMPLATE = [
    ("I. Details of the Listed Entity", [
        ("Corporate Identity Number (CIN)", "cin"),
        ("Name of the Listed Entity", "company_name"),
        ("Year of incorporation", "year_of_incorporation"),
        ("Registered office address", "registered_office"),
        ("Corporate address", "corporate_office"),
        ("E-mail", "email"),
        ("Telephone", "telephone"),
        ("Website", "website"),
        ("Financial year for which reporting is being done", "financial_year"),
        ("Name of the Stock Exchange(s)", "stock_exchange"),
        ("Paid-up Capital (INR)", "paid_up_capital"),
        ("Reporting boundary (Standalone/Consolidated)", "reporting_boundary"),
        ("Name of assurance provider", "assurance_provider"),
        ("Type of assurance obtained", "assurance_type"),
    ]),
    ("II. Products/Services", [
        ("Details of business activities (90% of turnover)", "business_activities"),
        ("Products/Services sold (90% of turnover)", "products_services_sold"),
    ]),
    ("III. Operations", [
        ("Number of plants - National", "plants_national"),
        ("Number of plants - International", "plants_international"),
        ("Number of offices - National", "offices_national"),
        ("Number of offices - International", "offices_international"),
        ("Markets served - States/UTs", "states_served"),
        ("Markets served - Countries", "countries_served"),
        ("Exports as % of total turnover", "export_pct"),
        ("Types of customers", "customer_types"),
    ]),
    ("IV. Employees", [
        ("Permanent Employees - Male", "employees_permanent_male"),
        ("Permanent Employees - Female", "employees_permanent_female"),
        ("Total Permanent Employees", "employees_permanent"),
        ("Other than Permanent - Male", "employees_other_male"),
        ("Other than Permanent - Female", "employees_other_female"),
        ("Permanent Workers - Male", "workers_permanent_male"),
        ("Permanent Workers - Female", "workers_permanent_female"),
        ("Women on Board of Directors (%)", "women_board_pct"),
        ("Women in KMP (%)", "women_kmp_pct"),
        ("Turnover rate - Permanent Employees (%)", "employee_turnover_rate"),
    ]),
    ("VI. CSR Details", [
        ("Whether CSR is applicable", "csr_applicable"),
        ("Turnover (INR)", "turnover"),
        ("Net worth (INR)", "net_worth"),
    ]),
]

SECTION_B_PRINCIPLES = [
    "P1: Ethics, Transparency and Accountability",
    "P2: Products and Services (Lifecycle Sustainability)",
    "P3: Employee Wellbeing",
    "P4: Stakeholder Engagement",
    "P5: Human Rights",
    "P6: Environment",
    "P7: Policy Advocacy",
    "P8: Inclusive Growth (CSR)",
    "P9: Consumer/Customer Value",
]

SECTION_C_TEMPLATE = [
    ("Principle 1: Ethics & Governance", [
        ("Code of Conduct / Ethics Policy", "code_of_conduct"),
        ("Anti-corruption / Anti-bribery policy", "anti_corruption_policy"),
        ("Complaints on ethics (number)", "complaints_ethics"),
    ]),
    ("Principle 2: Product Lifecycle Sustainability", [
        ("R&D / Capex for sustainability (INR)", "r_and_d_spend"),
        ("Sustainable sourcing (%)", "sustainable_sourcing_pct"),
        ("Recycled / reused input material (%)", "recycled_input_pct"),
    ]),
    ("Principle 3: Employee Wellbeing", [
        ("Employee turnover rate (%)", "employee_turnover_rate"),
        ("Safety incidents (LTIFR)", "safety_incidents"),
        ("Training hours per employee", "training_hours_per_employee"),
        ("Median salary - Male (INR)", "median_salary_male"),
        ("Median salary - Female (INR)", "median_salary_female"),
    ]),
    ("Principle 4: Stakeholder Engagement", [
        ("Key stakeholder groups identified", "stakeholder_groups_identified"),
    ]),
    ("Principle 5: Human Rights", [
        ("Human rights training (%)", "human_rights_training_pct"),
        ("Child labour complaints", "child_labor_complaints"),
    ]),
    ("Principle 6: Environment", [
        ("Total energy consumption (GJ)", "energy_consumption_total"),
        ("Renewable energy (%)", "renewable_energy_pct"),
        ("Total water withdrawal (KL)", "water_withdrawal"),
        ("GHG Scope 1 emissions (tCO2e)", "ghg_scope1"),
        ("GHG Scope 2 emissions (tCO2e)", "ghg_scope2"),
        ("Total waste generated (MT)", "waste_generated"),
        ("Waste recycled/reused (%)", "waste_recycled_pct"),
    ]),
    ("Principle 7: Policy Advocacy", [
        ("Trade/industry association memberships", "trade_associations"),
    ]),
    ("Principle 8: Inclusive Growth", [
        ("CSR spend (INR)", "csr_spend"),
        ("Community beneficiaries (number)", "community_beneficiaries"),
    ]),
    ("Principle 9: Consumer/Customer", [
        ("Consumer complaints (number)", "consumer_complaints"),
        ("Data privacy complaints", "data_privacy_complaints"),
        ("Product recalls (number)", "product_recalls"),
    ]),
]


# ═══════════════════════════════════════════════════════════════════════
# EXCEL GENERATOR
# ═══════════════════════════════════════════════════════════════════════


def _get_value(extracted_data: dict, field: str) -> str:
    """Get a field value from extracted data (any section)."""
    for section_data in extracted_data.values():
        if isinstance(section_data, dict) and field in section_data:
            val = section_data[field]
            if val is not None:
                return str(val)
    return ""


def generate_brsr_excel(extracted_data: dict, company_name: str = "Company") -> bytes:
    """
    Generate a SEBI BRSR format Excel workbook from extracted data.

    Returns: bytes of the .xlsx file
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")

    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=11)
    subheader_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    normal_font = Font(name="Calibri", size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ─── Sheet 1: Section A ───────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = "Section A - General"
    ws_a.column_dimensions["A"].width = 45
    ws_a.column_dimensions["B"].width = 50

    row = 1
    ws_a.cell(row=row, column=1, value="BUSINESS RESPONSIBILITY AND SUSTAINABILITY REPORT")
    ws_a.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
    row += 1
    ws_a.cell(row=row, column=1, value=f"Entity: {_get_value(extracted_data, 'company_name') or company_name}")
    ws_a.cell(row=row, column=2, value=f"FY: {_get_value(extracted_data, 'financial_year')}")
    row += 2

    for section_title, fields in SECTION_A_TEMPLATE:
        ws_a.cell(row=row, column=1, value=section_title)
        ws_a.cell(row=row, column=1).font = subheader_font
        ws_a.cell(row=row, column=1).fill = subheader_fill
        ws_a.cell(row=row, column=2).fill = subheader_fill
        row += 1
        for label, field_key in fields:
            ws_a.cell(row=row, column=1, value=label).font = normal_font
            ws_a.cell(row=row, column=1).border = border
            ws_a.cell(row=row, column=2, value=_get_value(extracted_data, field_key)).font = normal_font
            ws_a.cell(row=row, column=2).border = border
            row += 1
        row += 1

    # ─── Sheet 2: Section B ───────────────────────────────────────────
    ws_b = wb.create_sheet("Section B - Management")
    ws_b.column_dimensions["A"].width = 50
    ws_b.column_dimensions["B"].width = 20
    ws_b.column_dimensions["C"].width = 20
    ws_b.column_dimensions["D"].width = 30

    row = 1
    ws_b.cell(row=row, column=1, value="SECTION B: MANAGEMENT AND PROCESS DISCLOSURES")
    ws_b.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 2

    headers = ["Principle", "Policy Available (Y/N)", "Board Approved (Y/N)", "Web Link"]
    for col, h in enumerate(headers, 1):
        ws_b.cell(row=row, column=col, value=h)
        ws_b.cell(row=row, column=col).font = header_font
        ws_b.cell(row=row, column=col).fill = header_fill
        ws_b.cell(row=row, column=col).border = border
    row += 1

    policy_available = _get_value(extracted_data, "policy_available")
    policy_approved = _get_value(extracted_data, "policy_approved_by_board")
    policy_link = _get_value(extracted_data, "policy_web_link")

    for principle in SECTION_B_PRINCIPLES:
        ws_b.cell(row=row, column=1, value=principle).font = normal_font
        ws_b.cell(row=row, column=1).border = border
        ws_b.cell(row=row, column=2, value=policy_available).border = border
        ws_b.cell(row=row, column=3, value=policy_approved).border = border
        ws_b.cell(row=row, column=4, value=policy_link).border = border
        row += 1

    row += 2
    ws_b.cell(row=row, column=1, value="Grievance Redressal Mechanism")
    ws_b.cell(row=row, column=1).font = subheader_font
    row += 1
    ws_b.cell(row=row, column=1, value=_get_value(extracted_data, "grievance_mechanism"))

    # ─── Sheet 3: Section C ───────────────────────────────────────────
    ws_c = wb.create_sheet("Section C - Principles")
    ws_c.column_dimensions["A"].width = 45
    ws_c.column_dimensions["B"].width = 35
    ws_c.column_dimensions["C"].width = 15

    row = 1
    ws_c.cell(row=row, column=1, value="SECTION C: PRINCIPLE-WISE PERFORMANCE DISCLOSURE")
    ws_c.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12)
    row += 2

    # Headers
    for col, h in enumerate(["Disclosure", "Value", "Status"], 1):
        ws_c.cell(row=row, column=col, value=h)
        ws_c.cell(row=row, column=col).font = header_font
        ws_c.cell(row=row, column=col).fill = header_fill
        ws_c.cell(row=row, column=col).border = border
    row += 1

    for principle_title, fields in SECTION_C_TEMPLATE:
        ws_c.cell(row=row, column=1, value=principle_title)
        ws_c.cell(row=row, column=1).font = subheader_font
        ws_c.cell(row=row, column=1).fill = subheader_fill
        ws_c.cell(row=row, column=2).fill = subheader_fill
        ws_c.cell(row=row, column=3).fill = subheader_fill
        row += 1
        for label, field_key in fields:
            value = _get_value(extracted_data, field_key)
            status = "✓ Disclosed" if value else "✗ Missing"
            ws_c.cell(row=row, column=1, value=label).font = normal_font
            ws_c.cell(row=row, column=1).border = border
            ws_c.cell(row=row, column=2, value=value).font = normal_font
            ws_c.cell(row=row, column=2).border = border
            ws_c.cell(row=row, column=3, value=status).font = normal_font
            ws_c.cell(row=row, column=3).border = border
            row += 1
        row += 1

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_brsr_core_excel(extracted_data: dict) -> bytes:
    """
    Generate BRSR Core subset Excel — only the fields subject to reasonable assurance.
    Used by auditors / assurance providers.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required")

    wb = Workbook()
    ws = wb.active
    ws.title = "BRSR Core - Assurance"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    row = 1
    ws.cell(row=row, column=1, value="BRSR CORE — Indicators Subject to Assurance")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=13)
    row += 1
    ws.cell(row=row, column=1, value=f"Entity: {_get_value(extracted_data, 'company_name')}")
    ws.cell(row=row, column=2, value=f"FY: {_get_value(extracted_data, 'financial_year')}")
    row += 2

    headers = ["BRSR Core Indicator", "Extracted Value", "Status", "Assurance Phase"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).border = border
    row += 1

    core_indicators = [
        ("GHG Scope 1 emissions (tCO2e)", "ghg_scope1", "FY2024-25"),
        ("GHG Scope 2 emissions (tCO2e)", "ghg_scope2", "FY2024-25"),
        ("GHG emission intensity", "ghg_intensity", "FY2024-25"),
        ("Total energy consumption (GJ)", "energy_consumption_total", "FY2024-25"),
        ("Renewable energy share (%)", "renewable_energy_pct", "FY2024-25"),
        ("Water withdrawal (KL)", "water_withdrawal", "FY2024-25"),
        ("Water intensity", "water_intensity", "FY2024-25"),
        ("Waste generated (MT)", "waste_generated", "FY2025-26"),
        ("Waste recycled (%)", "waste_recycled_pct", "FY2025-26"),
        ("Women on Board (%)", "women_board_pct", "FY2024-25"),
        ("Employee turnover rate (%)", "employee_turnover_rate", "FY2024-25"),
        ("LTIFR / Safety incidents", "safety_incidents", "FY2024-25"),
        ("Training hours per employee", "training_hours_per_employee", "FY2025-26"),
        ("Median remuneration - Male", "median_salary_male", "FY2025-26"),
        ("Median remuneration - Female", "median_salary_female", "FY2025-26"),
    ]

    for label, field_key, phase in core_indicators:
        value = _get_value(extracted_data, field_key)
        status = "Disclosed" if value else "MISSING"
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=value or "—").border = border
        ws.cell(row=row, column=3, value=status).border = border
        ws.cell(row=row, column=4, value=phase).border = border
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
