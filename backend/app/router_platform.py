"""
Platform API Router - Data Entry, Carbon Calculator, Action Plans, Calendar, Reports.
Full Greenomy-equivalent API for Indian ESG compliance.
"""

from fastapi import APIRouter, HTTPException, Header, Depends
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import date, datetime
import json

from app.config import get_settings
from app.carbon_calculator import (
    calculate_scope1_emissions,
    calculate_scope2_emissions,
    calculate_scope3_emissions,
    calculate_energy_intensity,
    calculate_ghg_intensity,
    get_pat_compliance,
    CEA_GRID_EMISSION_FACTORS,
    SEBI_COMPLIANCE_CALENDAR,
    STATIONARY_COMBUSTION_FACTORS,
    MOBILE_COMBUSTION_FACTORS,
    SCOPE_3_FACTORS,
    PAT_SECTOR_BENCHMARKS,
)
from app.action_plan import generate_action_plan_from_gaps, NGRBC_PRINCIPLES
from app.brsr_datapoints import BRSR_DATAPOINTS, get_datapoints_stats

router = APIRouter(prefix="/api/platform", tags=["Platform"])

settings = get_settings()


def get_supabase_admin():
    from supabase import create_client
    return create_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_KEY)


async def verify_auth(authorization: str = Header(...)) -> str:
    """Verify JWT and return user_id."""
    expected_token = f"Bearer {settings.SUPABASE_SERVICE_KEY}"
    if authorization == expected_token:
        return "service_role"  # Internal service calls
    # For user JWTs, decode with Supabase
    # In production, verify JWT signature
    token = authorization.replace("Bearer ", "")
    if not token:
        raise HTTPException(status_code=401, detail="Missing auth token")
    return token  # Simplified; in prod, decode JWT to get user_id


# ═══════════════════════════════════════════════════════════════
# DATA ENTRY APIs
# ═══════════════════════════════════════════════════════════════

class DataEntryRequest(BaseModel):
    financial_year: str = Field(..., example="FY2024-25")
    datapoint_id: str = Field(..., example="A.I.1")
    value: dict | str | int | float | bool
    source: str = "manual"
    user_id: Optional[str] = None
    notes: Optional[str] = None


class BulkDataEntryRequest(BaseModel):
    financial_year: str
    user_id: Optional[str] = None
    entries: List[dict]  # [{datapoint_id, value, notes}]


@router.post("/data-entry")
async def save_data_entry(req: DataEntryRequest, authorization: str = Header(...)):
    """Save a single BRSR data point entry."""
    supabase = get_supabase_admin()
    
    # Find the datapoint metadata
    dp = next((d for d in BRSR_DATAPOINTS if d["id"] == req.datapoint_id), None)
    if not dp:
        raise HTTPException(status_code=400, detail=f"Invalid datapoint_id: {req.datapoint_id}")
    
    entry_data = {
        "financial_year": req.financial_year,
        "datapoint_id": req.datapoint_id,
        "section": dp["section"],
        "subsection": dp.get("subsection", ""),
        "value": json.dumps(req.value) if not isinstance(req.value, str) else req.value,
        "source": req.source,
        "notes": req.notes,
    }
    if req.user_id:
        entry_data["user_id"] = req.user_id
    
    # Upsert (update if exists, insert if not)
    result = supabase.table("brsr_entries").upsert(
        entry_data,
        on_conflict="user_id,financial_year,datapoint_id"
    ).execute()
    
    return {"status": "saved", "datapoint_id": req.datapoint_id, "entry": result.data}


@router.post("/data-entry/bulk")
async def save_bulk_entries(req: BulkDataEntryRequest, authorization: str = Header(...)):
    """Save multiple BRSR data points at once."""
    supabase = get_supabase_admin()
    saved = 0
    errors = []
    
    for entry in req.entries:
        dp_id = entry.get("datapoint_id")
        dp = next((d for d in BRSR_DATAPOINTS if d["id"] == dp_id), None)
        if not dp:
            errors.append(f"Invalid: {dp_id}")
            continue
        
        entry_data = {
            "financial_year": req.financial_year,
            "datapoint_id": dp_id,
            "section": dp["section"],
            "subsection": dp.get("subsection", ""),
            "value": json.dumps(entry.get("value", "")),
            "source": entry.get("source", "manual"),
            "notes": entry.get("notes"),
        }
        if req.user_id:
            entry_data["user_id"] = req.user_id
        
        try:
            supabase.table("brsr_entries").upsert(
                entry_data,
                on_conflict="user_id,financial_year,datapoint_id"
            ).execute()
            saved += 1
        except Exception as e:
            errors.append(f"{dp_id}: {str(e)}")
    
    return {"status": "completed", "saved": saved, "errors": errors}


@router.get("/data-entry/{financial_year}")
async def get_entries(financial_year: str, section: Optional[str] = None, user_id: Optional[str] = None, authorization: str = Header(...)):
    """Get all data entries for a financial year."""
    supabase = get_supabase_admin()
    
    query = supabase.table("brsr_entries").select("*").eq("financial_year", financial_year)
    if section:
        query = query.eq("section", section)
    if user_id:
        query = query.eq("user_id", user_id)
    
    result = query.order("datapoint_id").execute()
    
    # Calculate completion stats
    total_mandatory = len([d for d in BRSR_DATAPOINTS if d["mandatory"]])
    total_core = len([d for d in BRSR_DATAPOINTS if d.get("core")])
    filled_ids = {e["datapoint_id"] for e in (result.data or [])}
    filled_mandatory = len([d for d in BRSR_DATAPOINTS if d["mandatory"] and d["id"] in filled_ids])
    filled_core = len([d for d in BRSR_DATAPOINTS if d.get("core") and d["id"] in filled_ids])
    
    return {
        "financial_year": financial_year,
        "entries": result.data or [],
        "completion": {
            "total_mandatory": total_mandatory,
            "filled_mandatory": filled_mandatory,
            "mandatory_percent": round((filled_mandatory / total_mandatory) * 100, 1) if total_mandatory > 0 else 0,
            "total_core": total_core,
            "filled_core": filled_core,
            "core_percent": round((filled_core / total_core) * 100, 1) if total_core > 0 else 0,
        }
    }


@router.get("/data-entry/{financial_year}/progress")
async def get_progress(financial_year: str, authorization: str = Header(...)):
    """Get section-wise completion progress."""
    supabase = get_supabase_admin()
    
    result = supabase.table("brsr_entries").select("datapoint_id, section").eq("financial_year", financial_year).execute()
    filled_ids = {e["datapoint_id"] for e in (result.data or [])}
    
    sections = {}
    for dp in BRSR_DATAPOINTS:
        section = dp["section"]
        if section not in sections:
            sections[section] = {"total": 0, "filled": 0, "mandatory": 0, "mandatory_filled": 0}
        sections[section]["total"] += 1
        if dp["id"] in filled_ids:
            sections[section]["filled"] += 1
        if dp["mandatory"]:
            sections[section]["mandatory"] += 1
            if dp["id"] in filled_ids:
                sections[section]["mandatory_filled"] += 1
    
    # Calculate percentages
    for section in sections:
        s = sections[section]
        s["percent"] = round((s["filled"] / s["total"]) * 100, 1) if s["total"] > 0 else 0
        s["mandatory_percent"] = round((s["mandatory_filled"] / s["mandatory"]) * 100, 1) if s["mandatory"] > 0 else 0
    
    return {"financial_year": financial_year, "sections": sections}


# ═══════════════════════════════════════════════════════════════
# DOWNLOAD FILLED DATA AS EXCEL
# ═══════════════════════════════════════════════════════════════

@router.get("/data-entry/{financial_year}/download-excel")
async def download_data_entry_excel(financial_year: str, authorization: str = Header(...)):
    """Download all filled BRSR entries as an Excel file."""
    from fastapi.responses import Response
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    supabase = get_supabase_admin()
    result = supabase.table("brsr_entries").select("*").eq("financial_year", financial_year).order("datapoint_id").execute()
    entries = result.data or []

    if not entries:
        raise HTTPException(status_code=404, detail="No entries found for this financial year")

    entry_map = {e["datapoint_id"]: e for e in entries}

    wb = Workbook()
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
    mandatory_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    filled_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    section_dps: dict = {"section_a": [], "section_b": [], "section_c": []}
    for dp in BRSR_DATAPOINTS:
        section_dps.setdefault(dp["section"], []).append(dp)

    for idx, (section_key, dps) in enumerate(section_dps.items()):
        if idx == 0:
            ws = wb.active
            ws.title = section_key.replace("_", " ").title()
        else:
            ws = wb.create_sheet(section_key.replace("_", " ").title())

        headers = ["Datapoint ID", "Label", "Mandatory", "Core", "Value", "Source", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, dp in enumerate(dps, 2):
            entry = entry_map.get(dp["id"])
            value = entry["value"] if entry else ""
            source = entry.get("source", "") if entry else ""
            status = "Filled" if entry else "Missing"

            ws.cell(row=row_idx, column=1, value=dp["id"]).border = border
            ws.cell(row=row_idx, column=2, value=dp["label"]).border = border
            ws.cell(row=row_idx, column=3, value="Yes" if dp["mandatory"] else "No").border = border
            ws.cell(row=row_idx, column=4, value="Yes" if dp.get("core") else "No").border = border

            val_cell = ws.cell(row=row_idx, column=5, value=str(value) if value else "")
            val_cell.border = border
            if entry:
                val_cell.fill = filled_fill
            elif dp["mandatory"]:
                val_cell.fill = mandatory_fill

            ws.cell(row=row_idx, column=6, value=source).border = border
            ws.cell(row=row_idx, column=7, value=status).border = border

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BRSR_Data_Entry_{financial_year}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ═══════════════════════════════════════════════════════════════
# SAVE ALL ENTRIES (bulk across all sections)
# ═══════════════════════════════════════════════════════════════

@router.post("/data-entry/save-all")
async def save_all_entries(request: BulkDataEntryRequest, authorization: str = Header(...)):
    """Save all filled entries across all sections at once."""
    supabase = get_supabase_admin()
    saved = 0
    errors = []

    for entry in request.entries:
        try:
            dp_id = entry["datapoint_id"]
            dp = next((d for d in BRSR_DATAPOINTS if d["id"] == dp_id), None)
            section = dp["section"] if dp else ("section_a" if dp_id.startswith("A") else "section_b" if dp_id.startswith("B") else "section_c")
            subsection = dp.get("subsection", "") if dp else ""
            
            upsert_data = {
                "datapoint_id": dp_id,
                "financial_year": request.financial_year,
                "value": str(entry.get("value", "")),
                "source": entry.get("source", "manual"),
                "section": section,
                "subsection": subsection,
            }
            if request.user_id:
                upsert_data["user_id"] = request.user_id
            
            supabase.table("brsr_entries").upsert(
                upsert_data,
                on_conflict="user_id,financial_year,datapoint_id"
            ).execute()
            saved += 1
        except Exception as e:
            errors.append({"datapoint_id": entry["datapoint_id"], "error": str(e)})

    return {"saved": saved, "total": len(request.entries), "errors": errors[:5]}


# ═══════════════════════════════════════════════════════════════
# IMPORT FROM EXTRACTION
# ═══════════════════════════════════════════════════════════════

@router.post("/data-entry/import-extraction/{report_id}")
async def import_extraction_to_entries(report_id: str, authorization: str = Header(...)):
    """Import extracted data from a report into brsr_entries for data entry auto-fill."""
    from app.brsr_datapoints import _FIELD_TO_DATAPOINT_MAP
    
    supabase = get_supabase_admin()
    
    # Fetch the report's extracted data
    try:
        result = supabase.table("reports").select("extracted_data, financial_year, user_id").eq("id", report_id).single().execute()
    except Exception:
        raise HTTPException(status_code=404, detail="Report not found")
    if not result.data:
        raise HTTPException(status_code=404, detail="Report not found")
    
    report = result.data
    extracted_data = report.get("extracted_data", {})
    financial_year = report.get("financial_year") or "FY2024-25"
    user_id = report.get("user_id")
    
    if not extracted_data:
        raise HTTPException(status_code=400, detail="No extracted data in this report")
    
    if not user_id:
        raise HTTPException(status_code=400, detail="Report has no associated user")
    
    # Flatten all sections into a single dict of field_name -> value
    flat_data = {}
    for section_key, section_val in extracted_data.items():
        if isinstance(section_val, dict):
            for k, v in section_val.items():
                if v is not None and v != "" and v != "N/A" and k not in ("gap_analysis", "datapoints_stats", "benchmark"):
                    flat_data[k] = v
    
    # Map extracted fields to datapoint IDs
    imported = 0
    errors = []
    entries_to_upsert = []
    
    for field_name, value in flat_data.items():
        if field_name in _FIELD_TO_DATAPOINT_MAP:
            datapoint_ids = _FIELD_TO_DATAPOINT_MAP[field_name]
            for dp_id in datapoint_ids:
                dp = next((d for d in BRSR_DATAPOINTS if d["id"] == dp_id), None)
                if not dp:
                    continue
                entries_to_upsert.append({
                    "user_id": user_id,
                    "financial_year": financial_year,
                    "datapoint_id": dp_id,
                    "section": dp["section"],
                    "subsection": dp.get("subsection", ""),
                    "value": json.dumps(value) if not isinstance(value, str) else value,
                    "source": "ai_extracted",
                    "notes": f"Auto-imported from report {report_id}",
                })
    
    # Batch upsert
    for entry in entries_to_upsert:
        try:
            supabase.table("brsr_entries").upsert(
                entry,
                on_conflict="user_id,financial_year,datapoint_id"
            ).execute()
            imported += 1
        except Exception as e:
            errors.append(f"{entry['datapoint_id']}: {str(e)}")
    
    return {
        "status": "completed",
        "imported": imported,
        "total_fields_found": len(flat_data),
        "errors": errors[:10],
        "financial_year": financial_year,
    }


# ═══════════════════════════════════════════════════════════════
# MULTI-YEAR TRACKING
# ═══════════════════════════════════════════════════════════════

@router.get("/tracking/multi-year")
async def get_multi_year_data(authorization: str = Header(...)):
    """Get year-over-year progress for all financial years."""
    supabase = get_supabase_admin()
    
    result = supabase.table("brsr_entries").select("financial_year, datapoint_id, section, value").execute()
    
    years = {}
    for entry in (result.data or []):
        fy = entry["financial_year"]
        if fy not in years:
            years[fy] = {"total_entries": 0, "sections": {"section_a": 0, "section_b": 0, "section_c": 0}}
        years[fy]["total_entries"] += 1
        section = entry.get("section", "section_a")
        if section in years[fy]["sections"]:
            years[fy]["sections"][section] += 1
    
    # Calculate completion for each year
    total_datapoints = len(BRSR_DATAPOINTS)
    for fy in years:
        years[fy]["completion_percent"] = round((years[fy]["total_entries"] / total_datapoints) * 100, 1)
    
    return {"years": years, "total_datapoints": total_datapoints}


@router.get("/tracking/trends/{datapoint_id}")
async def get_datapoint_trend(datapoint_id: str, authorization: str = Header(...)):
    """Get historical values for a specific datapoint across years."""
    supabase = get_supabase_admin()
    
    result = supabase.table("brsr_entries").select("financial_year, value, source, updated_at").eq(
        "datapoint_id", datapoint_id
    ).order("financial_year").execute()
    
    dp = next((d for d in BRSR_DATAPOINTS if d["id"] == datapoint_id), None)
    
    return {
        "datapoint_id": datapoint_id,
        "metadata": dp,
        "trend": result.data or [],
    }


# ═══════════════════════════════════════════════════════════════
# CARBON CALCULATOR APIs
# ═══════════════════════════════════════════════════════════════

class Scope1Request(BaseModel):
    fuel_type: str
    quantity: float
    financial_year: Optional[str] = "FY2024-25"


class Scope2Request(BaseModel):
    electricity_mwh: float
    financial_year: str = "FY2024-25"
    state: str = "national"


class Scope3Request(BaseModel):
    category: str
    quantity: float


class CarbonSummaryRequest(BaseModel):
    financial_year: str = "FY2024-25"
    scope1_entries: List[dict] = []  # [{fuel_type, quantity}]
    scope2_entries: List[dict] = []  # [{electricity_mwh, state}]
    scope3_entries: List[dict] = []  # [{category, quantity}]
    revenue_crores: Optional[float] = None


@router.post("/carbon/scope1")
async def calc_scope1(req: Scope1Request):
    """Calculate Scope 1 emissions from fuel combustion."""
    result = calculate_scope1_emissions(req.fuel_type, req.quantity)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.post("/carbon/scope2")
async def calc_scope2(req: Scope2Request):
    """Calculate Scope 2 emissions from purchased electricity."""
    return calculate_scope2_emissions(req.electricity_mwh, req.financial_year, req.state)


@router.post("/carbon/scope3")
async def calc_scope3(req: Scope3Request):
    """Calculate Scope 3 emissions by category."""
    result = calculate_scope3_emissions(req.category, req.quantity)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.post("/carbon/summary")
async def carbon_summary(req: CarbonSummaryRequest):
    """Calculate complete carbon footprint summary."""
    scope1_total = 0
    scope1_details = []
    for entry in req.scope1_entries:
        result = calculate_scope1_emissions(entry["fuel_type"], entry["quantity"])
        if "error" not in result:
            scope1_total += result["total_tco2e"]
            scope1_details.append(result)
    
    scope2_total = 0
    scope2_details = []
    for entry in req.scope2_entries:
        result = calculate_scope2_emissions(
            entry["electricity_mwh"],
            req.financial_year,
            entry.get("state", "national")
        )
        scope2_total += result["total_tco2e"]
        scope2_details.append(result)
    
    scope3_total = 0
    scope3_details = []
    for entry in req.scope3_entries:
        result = calculate_scope3_emissions(entry["category"], entry["quantity"])
        if "error" not in result:
            scope3_total += result["total_tco2e"]
            scope3_details.append(result)
    
    total_emissions = scope1_total + scope2_total + scope3_total
    
    summary = {
        "financial_year": req.financial_year,
        "scope_1": {"total_tco2e": round(scope1_total, 2), "details": scope1_details},
        "scope_2": {"total_tco2e": round(scope2_total, 2), "details": scope2_details},
        "scope_3": {"total_tco2e": round(scope3_total, 2), "details": scope3_details},
        "total_emissions_tco2e": round(total_emissions, 2),
    }
    
    if req.revenue_crores and req.revenue_crores > 0:
        summary["ghg_intensity"] = calculate_ghg_intensity(total_emissions, req.revenue_crores)
    
    # BRSR disclosure mapping
    summary["brsr_mapping"] = {
        "C.P6.GHG.1": round(scope1_total, 2),
        "C.P6.GHG.2": round(scope2_total, 2),
        "C.P6.GHG.3": round(scope3_total, 2),
        "total_scope_1_2": round(scope1_total + scope2_total, 2),
    }
    
    return summary


@router.get("/carbon/factors")
async def get_emission_factors():
    """Get all available emission factors for reference."""
    return {
        "stationary_combustion": {k: {"unit": v["unit"], "total_factor": v["co2"] + v["ch4"]*28 + v["n2o"]*265} 
                                  for k, v in STATIONARY_COMBUSTION_FACTORS.items()},
        "mobile_combustion": {k: {"unit": v["unit"], "total_factor": v["co2"] + v["ch4"]*28 + v["n2o"]*265} 
                             for k, v in MOBILE_COMBUSTION_FACTORS.items()},
        "scope_3_categories": {k: {"unit": v["unit"], "factor": v["factor"], "source": v["source"]} 
                              for k, v in SCOPE_3_FACTORS.items()},
        "grid_factors": CEA_GRID_EMISSION_FACTORS,
        "pat_sectors": list(PAT_SECTOR_BENCHMARKS.keys()),
    }


@router.post("/carbon/pat-check")
async def check_pat_compliance(sector: str, actual_sec: float):
    """Check PAT scheme compliance for designated consumers."""
    result = get_pat_compliance(sector, actual_sec)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


# ═══════════════════════════════════════════════════════════════
# ACTION PLAN APIs
# ═══════════════════════════════════════════════════════════════

class ActionPlanGenerateRequest(BaseModel):
    financial_year: str = "FY2024-25"
    report_id: Optional[str] = None
    sector: str = "general"


@router.post("/action-plan/generate")
async def generate_action_plan(req: ActionPlanGenerateRequest, authorization: str = Header(...)):
    """Generate AI-powered action plan from gap analysis."""
    supabase = get_supabase_admin()
    
    # Get gap analysis data
    if req.report_id:
        report = supabase.table("reports").select("extracted_data").eq("id", req.report_id).single().execute()
        if not report.data:
            raise HTTPException(status_code=404, detail="Report not found")
        
        extracted_data = report.data.get("extracted_data", {})
        gap_analysis = extracted_data.get("gap_analysis", {})
    else:
        # Generate from manual entries
        entries = supabase.table("brsr_entries").select("*").eq("financial_year", req.financial_year).execute()
        filled_ids = {e["datapoint_id"] for e in (entries.data or [])}
        
        missing_essential = [d for d in BRSR_DATAPOINTS if d["mandatory"] and d["id"] not in filled_ids]
        missing_leadership = [d for d in BRSR_DATAPOINTS if d.get("indicator_type") == "leadership" and d["id"] not in filled_ids]
        
        gap_analysis = {
            "total_datapoints": len(BRSR_DATAPOINTS),
            "filled_datapoints": len(filled_ids),
            "missing_essential": missing_essential,
            "missing_essential_count": len(missing_essential),
            "missing_leadership": missing_leadership,
            "missing_leadership_count": len(missing_leadership),
            "completion_percent": round((len(filled_ids) / len(BRSR_DATAPOINTS)) * 100, 1),
        }
        extracted_data = {}
    
    # Generate action plan
    actions = generate_action_plan_from_gaps(gap_analysis, extracted_data, req.sector)
    
    return {
        "financial_year": req.financial_year,
        "sector": req.sector,
        "total_actions": len(actions),
        "by_priority": {
            "critical": len([a for a in actions if a["priority"] == "critical"]),
            "high": len([a for a in actions if a["priority"] == "high"]),
            "medium": len([a for a in actions if a["priority"] == "medium"]),
            "low": len([a for a in actions if a["priority"] == "low"]),
        },
        "estimated_total_cost_inr": sum(a.get("estimated_cost_inr", 0) for a in actions),
        "actions": actions,
    }


@router.get("/action-plan/{financial_year}")
async def get_action_plans(financial_year: str, authorization: str = Header(...)):
    """Get saved action plans for a financial year."""
    supabase = get_supabase_admin()
    
    result = supabase.table("action_plans").select("*").eq("financial_year", financial_year).order("priority").execute()
    
    return {"financial_year": financial_year, "plans": result.data or []}


class UpdateActionPlanRequest(BaseModel):
    status: str
    completion_notes: Optional[str] = None


@router.patch("/action-plan/{plan_id}")
async def update_action_plan(plan_id: str, req: UpdateActionPlanRequest, authorization: str = Header(...)):
    """Update action plan status."""
    supabase = get_supabase_admin()
    
    update_data = {"status": req.status}
    if req.status == "completed":
        update_data["completed_at"] = datetime.utcnow().isoformat()
        update_data["completion_notes"] = req.completion_notes
    
    result = supabase.table("action_plans").update(update_data).eq("id", plan_id).execute()
    return {"status": "updated", "plan": result.data}


# ═══════════════════════════════════════════════════════════════
# COMPLIANCE CALENDAR APIs
# ═══════════════════════════════════════════════════════════════

@router.get("/calendar/events")
async def get_calendar_events(financial_year: Optional[str] = None, authorization: str = Header(...)):
    """Get compliance calendar events."""
    supabase = get_supabase_admin()
    
    query = supabase.table("compliance_events").select("*").order("due_date")
    if financial_year:
        query = query.eq("financial_year", financial_year)
    
    result = query.execute()
    
    return {"events": result.data or []}


@router.get("/calendar/sebi-deadlines")
async def get_sebi_deadlines(financial_year: str = "FY2024-25"):
    """Get standard SEBI compliance deadlines for a financial year."""
    # Parse FY to get dates
    # FY2024-25 → April 2024 to March 2025, deadlines in Q2/Q3 FY next year
    fy_start_year = int(financial_year.replace("FY", "").split("-")[0]) + 2000
    
    deadlines = []
    for event in SEBI_COMPLIANCE_CALENDAR:
        if event["month"]:
            # Filing deadlines are typically after FY end
            deadline_year = fy_start_year + 1 if event["month"] >= 4 else fy_start_year + 2
            due_date = f"{deadline_year}-{event['month']:02d}-{event['day']:02d}"
        else:
            due_date = None
        
        deadlines.append({
            "title": event["title"],
            "description": event["description"],
            "regulatory_body": event["regulatory_body"],
            "due_date": due_date,
            "recurring": event["recurring"],
            "applies_to": event["applies_to"],
        })
    
    return {"financial_year": financial_year, "deadlines": deadlines}


class CreateCalendarEventRequest(BaseModel):
    title: str
    description: Optional[str] = None
    event_type: str = "custom"
    due_date: str  # ISO date
    financial_year: Optional[str] = None
    reminder_days: List[int] = [30, 7, 1]


@router.post("/calendar/events")
async def create_calendar_event(req: CreateCalendarEventRequest, authorization: str = Header(...)):
    """Create a custom compliance calendar event."""
    supabase = get_supabase_admin()
    
    event_data = {
        "title": req.title,
        "description": req.description,
        "event_type": req.event_type,
        "due_date": req.due_date,
        "financial_year": req.financial_year,
        "reminder_days": req.reminder_days,
    }
    
    result = supabase.table("compliance_events").insert(event_data).execute()
    return {"status": "created", "event": result.data}


# ═══════════════════════════════════════════════════════════════
# REPORT GENERATION APIs
# ═══════════════════════════════════════════════════════════════

class GenerateReportRequest(BaseModel):
    financial_year: str
    report_type: str = "brsr_full"  # brsr_full, brsr_core, brsr_lite, gap_analysis
    format: str = "pdf"  # pdf, docx, xlsx


@router.post("/reports/generate")
async def generate_report(req: GenerateReportRequest, authorization: str = Header(...)):
    """Generate a BRSR format report from collected data."""
    supabase = get_supabase_admin()
    
    # Get all entries for the financial year
    entries_result = supabase.table("brsr_entries").select("*").eq("financial_year", req.financial_year).execute()
    entries = entries_result.data or []
    
    if not entries:
        raise HTTPException(status_code=400, detail="No data entries found for this financial year. Please enter data first.")
    
    # Organize entries by section/subsection
    organized = {"section_a": {}, "section_b": {}, "section_c": {}}
    for entry in entries:
        section = entry["section"]
        if section in organized:
            organized[section][entry["datapoint_id"]] = {
                "value": entry["value"],
                "source": entry["source"],
                "verified": entry.get("verified", False),
            }
    
    # Build BRSR report structure
    report_content = {
        "report_type": req.report_type,
        "financial_year": req.financial_year,
        "generated_at": datetime.utcnow().isoformat(),
        "data": organized,
        "statistics": {
            "total_entries": len(entries),
            "by_section": {
                "section_a": len(organized["section_a"]),
                "section_b": len(organized["section_b"]),
                "section_c": len(organized["section_c"]),
            },
            "verified_entries": len([e for e in entries if e.get("verified")]),
            "ai_extracted": len([e for e in entries if e.get("source") == "ai_extracted"]),
            "manual": len([e for e in entries if e.get("source") == "manual"]),
        },
        "compliance_status": _get_compliance_status(organized, req.report_type),
    }
    
    # Save generated report metadata
    report_record = {
        "financial_year": req.financial_year,
        "report_type": req.report_type,
        "title": f"BRSR Report {req.financial_year}",
        "content": report_content,
        "format": req.format,
        "status": "draft",
    }
    
    save_result = supabase.table("generated_reports").insert(report_record).execute()
    
    return {
        "status": "generated",
        "report_id": save_result.data[0]["id"] if save_result.data else None,
        "report": report_content,
    }


@router.get("/reports/{financial_year}")
async def get_generated_reports(financial_year: str, authorization: str = Header(...)):
    """Get all generated reports for a financial year."""
    supabase = get_supabase_admin()
    
    result = supabase.table("generated_reports").select("*").eq(
        "financial_year", financial_year
    ).order("created_at", desc=True).execute()
    
    return {"reports": result.data or []}


# ═══════════════════════════════════════════════════════════════
# MATERIALITY ASSESSMENT
# ═══════════════════════════════════════════════════════════════

@router.get("/materiality/{financial_year}")
async def get_materiality(financial_year: str, authorization: str = Header(...)):
    """Get materiality assessment for a financial year."""
    supabase = get_supabase_admin()
    
    result = supabase.table("materiality_topics").select("*").eq("financial_year", financial_year).execute()
    return {"financial_year": financial_year, "topics": result.data or []}


class MaterialityTopicRequest(BaseModel):
    financial_year: str
    topic: str
    category: str  # environmental, social, governance
    impact_significance: float = Field(..., ge=0, le=1)
    financial_significance: float = Field(..., ge=0, le=1)
    stakeholder_relevance: float = Field(..., ge=0, le=1)
    brsr_principles: List[str] = []
    description: Optional[str] = None


@router.post("/materiality/topics")
async def add_materiality_topic(req: MaterialityTopicRequest, authorization: str = Header(...)):
    """Add a materiality topic assessment."""
    supabase = get_supabase_admin()
    
    topic_data = {
        "financial_year": req.financial_year,
        "topic": req.topic,
        "category": req.category,
        "impact_significance": req.impact_significance,
        "financial_significance": req.financial_significance,
        "stakeholder_relevance": req.stakeholder_relevance,
        "brsr_principles": req.brsr_principles,
        "description": req.description,
    }
    
    result = supabase.table("materiality_topics").insert(topic_data).execute()
    return {"status": "created", "topic": result.data}


# ═══════════════════════════════════════════════════════════════
# DATAPOINTS REFERENCE
# ═══════════════════════════════════════════════════════════════

@router.get("/datapoints")
async def get_all_datapoints(section: Optional[str] = None, mandatory_only: bool = False, core_only: bool = False):
    """Get BRSR datapoints reference list."""
    filtered = BRSR_DATAPOINTS
    
    if section:
        filtered = [d for d in filtered if d["section"] == section]
    if mandatory_only:
        filtered = [d for d in filtered if d["mandatory"]]
    if core_only:
        filtered = [d for d in filtered if d.get("core")]
    
    return {
        "total": len(filtered),
        "datapoints": filtered,
    }


@router.get("/datapoints/by-principle/{principle}")
async def get_datapoints_by_principle(principle: str):
    """Get datapoints for a specific NGRBC principle (P1-P9)."""
    principle_upper = principle.upper()
    filtered = [d for d in BRSR_DATAPOINTS if d.get("id", "").startswith(f"C.{principle_upper}")]
    
    return {
        "principle": principle_upper,
        "name": NGRBC_PRINCIPLES.get(principle_upper, {}).get("name", ""),
        "focus": NGRBC_PRINCIPLES.get(principle_upper, {}).get("focus", ""),
        "total": len(filtered),
        "datapoints": filtered,
    }


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def _get_compliance_status(organized: dict, report_type: str) -> dict:
    """Calculate compliance status for a report."""
    if report_type == "brsr_core":
        relevant = [d for d in BRSR_DATAPOINTS if d.get("core")]
    elif report_type == "brsr_lite":
        # BRSR Lite has ~40 essential indicators
        relevant = [d for d in BRSR_DATAPOINTS if d["mandatory"] and d.get("indicator_type") == "essential"][:40]
    else:
        relevant = [d for d in BRSR_DATAPOINTS if d["mandatory"]]
    
    filled_ids = set()
    for section_data in organized.values():
        filled_ids.update(section_data.keys())
    
    covered = [d for d in relevant if d["id"] in filled_ids]
    missing = [d for d in relevant if d["id"] not in filled_ids]
    
    return {
        "report_type": report_type,
        "required_datapoints": len(relevant),
        "covered_datapoints": len(covered),
        "missing_datapoints": len(missing),
        "compliance_percent": round((len(covered) / len(relevant)) * 100, 1) if relevant else 0,
        "status": "compliant" if len(missing) == 0 else "gaps_exist",
        "missing_ids": [d["id"] for d in missing[:20]],  # Top 20 missing
    }


# ═══════════════════════════════════════════════════════════════
# BOARD DASHBOARD API
# ═══════════════════════════════════════════════════════════════

@router.get("/board/dashboard")
async def board_dashboard(financial_year: str = "FY2025-26", authorization: str = Header(None)):
    """Executive board dashboard with compliance overview."""
    from supabase import create_client as create_supabase_client
    settings = get_settings()
    supabase = create_supabase_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_KEY)
    
    user_id = authorization.replace("Bearer ", "") if authorization else None
    if not user_id:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Get all entries for this user/FY
    entries_resp = supabase.table("brsr_entries").select("datapoint_id, value, verified").eq(
        "user_id", user_id
    ).eq("financial_year", financial_year).execute()
    entries = entries_resp.data or []
    
    # Get reports for this user
    reports_resp = supabase.table("reports").select("id, status, created_at, company_name").eq(
        "user_id", user_id
    ).execute()
    reports = reports_resp.data or []
    
    # Calculate completion
    total_required = 216  # BRSR Full has 216 datapoints
    filled = len([e for e in entries if e.get("value") and not e["datapoint_id"].startswith("CARBON_")])
    verified = len([e for e in entries if e.get("verified")])
    ai_extracted = len([r for r in reports if r.get("status") == "completed"])
    
    completion_pct = round((filled / total_required) * 100, 1) if total_required > 0 else 0
    verification_pct = round((verified / max(filled, 1)) * 100, 1)
    
    # Section progress (A: 1-60, B: 61-120, C: 121-216)
    section_a_total, section_b_total, section_c_total = 60, 60, 96
    section_a_filled = len([e for e in entries if e.get("datapoint_id", "").startswith("A.")])
    section_b_filled = len([e for e in entries if e.get("datapoint_id", "").startswith("B.")])
    section_c_filled = len([e for e in entries if e.get("datapoint_id", "").startswith("C.")])
    
    # Compliance score based on mandatory fields
    mandatory_datapoints = [d for d in BRSR_DATAPOINTS if d.get("mandatory")]
    mandatory_filled = len([e for e in entries if any(d["id"] == e["datapoint_id"] for d in mandatory_datapoints)])
    compliance_score = round((mandatory_filled / max(len(mandatory_datapoints), 1)) * 100, 1)
    
    # Upcoming deadlines from SEBI calendar
    from datetime import datetime as dt
    upcoming = []
    for item in SEBI_COMPLIANCE_CALENDAR:
        try:
            due = dt.strptime(item["due_date"], "%Y-%m-%d")
            today = dt.now()
            if due >= today:
                status = "upcoming" if (due - today).days > 30 else "due_soon"
                upcoming.append({
                    "regulation": item["regulation"],
                    "due_date": item["due_date"],
                    "status": status,
                })
        except:
            pass
    upcoming = sorted(upcoming, key=lambda x: x["due_date"])[:5]
    
    # Filing readiness
    blockers = []
    if completion_pct < 80:
        blockers.append(f"Only {completion_pct}% of datapoints filled (need ≥80%)")
    if verification_pct < 50:
        blockers.append(f"Only {verification_pct}% verified (recommended ≥50%)")
    if section_c_filled < 20:
        blockers.append("Section C (Principles) needs more disclosure")
    
    return {
        "financial_year": financial_year,
        "compliance_score": compliance_score,
        "completion": {
            "total_required": total_required,
            "filled": filled,
            "verified": verified,
            "ai_extracted": ai_extracted,
            "manual": max(filled - ai_extracted, 0),
            "completion_pct": completion_pct,
            "verification_pct": verification_pct,
        },
        "section_progress": {
            "section_a": {"filled": section_a_filled, "total": section_a_total, "pct": round((section_a_filled / section_a_total) * 100, 1)},
            "section_b": {"filled": section_b_filled, "total": section_b_total, "pct": round((section_b_filled / section_b_total) * 100, 1)},
            "section_c": {"filled": section_c_filled, "total": section_c_total, "pct": round((section_c_filled / section_c_total) * 100, 1)},
        },
        "risks": {
            "high_risk_suppliers": 0,
            "total_suppliers": 0,
            "non_compliant_regulations": len(mandatory_datapoints) - mandatory_filled,
            "overdue_filings": 0,
            "pending_approvals": 0,
        },
        "yoy": {
            "previous_year": "FY2024-25" if financial_year == "FY2025-26" else "FY2023-24",
            "prev_filled": 0,
            "disclosure_improvement": filled,
            "improvement_pct": None,
        },
        "esg_ratings": [
            {"agency": "MSCI ESG", "readiness_score": min(compliance_score + 10, 100)},
            {"agency": "Sustainalytics", "readiness_score": compliance_score},
            {"agency": "CDP Climate", "readiness_score": max(compliance_score - 5, 0)},
        ],
        "upcoming_deadlines": upcoming,
        "filing_readiness": {
            "ready": len(blockers) == 0,
            "blockers": blockers,
        },
    }
