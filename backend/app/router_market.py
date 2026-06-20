"""
Market-Readiness Router — Email Notifications, SEBI PDF Report, Bulk Excel Import,
YoY Comparative Analysis, Board-level Executive Dashboard.
"""

from fastapi import APIRouter, HTTPException, Header, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, date
import io
import json

from app.config import get_settings
from app.email_service import (
    send_team_invite, send_filing_reminder, send_extraction_complete,
    send_workflow_notification, send_deadline_alert, send_email, ADMIN_EMAIL,
)

router = APIRouter(prefix="/api/platform", tags=["Market Readiness"])
settings = get_settings()


def get_supabase_admin():
    from supabase import create_client
    return create_client(settings.SUPABASE_URL, settings.SUPABASE_SERVICE_KEY)


async def get_user_id(authorization: str) -> str:
    token = authorization.replace("Bearer ", "")
    if not token:
        raise HTTPException(status_code=401, detail="Missing auth token")
    try:
        import jwt as pyjwt
        payload = pyjwt.decode(token, options={"verify_signature": False})
        return payload.get("sub", token)
    except Exception:
        return token


# ═══════════════════════════════════════════════════════════════
# EMAIL NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════

class EmailNotification(BaseModel):
    template: str
    to_email: str
    variables: dict


@router.post("/notifications/send")
async def send_notification(body: EmailNotification, authorization: str = Header(...)):
    """Send an email notification using a template."""
    await get_user_id(authorization)
    result = await send_email(body.to_email, body.template, body.variables)
    return result


class FilingReminderRequest(BaseModel):
    financial_year: str = "FY2024-25"
    deadline_date: str = "2025-09-30"
    exchange: str = "BSE + NSE"


@router.post("/notifications/filing-reminder")
async def trigger_filing_reminder(body: FilingReminderRequest, authorization: str = Header(...)):
    """Send filing deadline reminders to all org members."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    # Get org members
    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    if not org_member.data:
        raise HTTPException(status_code=400, detail="No organization found")

    org_id = org_member.data["org_id"]
    members = sb.table("org_members").select("user_id").eq("org_id", org_id).execute()

    # Calculate completion
    entries = sb.table("brsr_entries").select("id").eq("org_id", org_id).eq("financial_year", body.financial_year).execute()
    completion_pct = min(int((len(entries.data or []) / 216) * 100), 100)

    # Calculate days remaining
    from datetime import datetime as dt
    try:
        deadline = dt.strptime(body.deadline_date, "%Y-%m-%d")
        days_remaining = max(0, (deadline - dt.now()).days)
    except ValueError:
        days_remaining = 30

    sent_count = 0
    for member in (members.data or []):
        profile = sb.table("profiles").select("email").eq("id", member["user_id"]).single().execute()
        if profile.data and profile.data.get("email"):
            await send_filing_reminder(
                to_email=profile.data["email"],
                financial_year=body.financial_year,
                deadline_date=body.deadline_date,
                days_remaining=days_remaining,
                completion_pct=completion_pct,
                pending_review=0,
                exchange=body.exchange,
            )
            sent_count += 1

    return {"status": "sent", "recipients": sent_count, "days_remaining": days_remaining}


# ═══════════════════════════════════════════════════════════════
# PDF REPORT GENERATION (SEBI BRSR FORMAT)
# ═══════════════════════════════════════════════════════════════

@router.get("/reports/brsr-pdf")
async def generate_brsr_report_pdf(
    authorization: str = Header(...),
    financial_year: str = Query(default="FY2024-25"),
    report_type: str = Query(default="brsr_full"),  # brsr_full, brsr_core, brsr_lite
):
    """Generate SEBI-prescribed BRSR format PDF report from user's data entries."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    # Get user's org and profile
    profile = sb.table("profiles").select("*").eq("id", user_id).single().execute()
    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    # Get organization details
    org_data = None
    if org_id:
        org_data = sb.table("organizations").select("*").eq("id", org_id).single().execute()

    # Get all BRSR entries for this FY
    query = sb.table("brsr_entries").select("*").eq("financial_year", financial_year)
    if org_id:
        query = query.eq("org_id", org_id)
    else:
        query = query.eq("user_id", user_id)
    entries = query.execute()

    # Get previous year entries for YoY comparison
    prev_fy = _get_previous_fy(financial_year)
    prev_query = sb.table("brsr_entries").select("*").eq("financial_year", prev_fy)
    if org_id:
        prev_query = prev_query.eq("org_id", org_id)
    else:
        prev_query = prev_query.eq("user_id", user_id)
    prev_entries = prev_query.execute()

    # Generate PDF
    from app.brsr_pdf_report import generate_sebi_brsr_pdf

    pdf_bytes = generate_sebi_brsr_pdf(
        entries=entries.data or [],
        prev_entries=prev_entries.data or [],
        company_name=org_data.data.get("name", profile.data.get("company_name", "Company")) if org_data and org_data.data else profile.data.get("company_name", "Company"),
        cin=org_data.data.get("cin", "") if org_data and org_data.data else "",
        financial_year=financial_year,
        report_type=report_type,
    )

    filename = f"BRSR_{report_type}_{financial_year}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
# BULK IMPORT (Excel/CSV)
# ═══════════════════════════════════════════════════════════════

@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    financial_year: str = Form(default="FY2024-25"),
    authorization: str = Header(...),
):
    """Bulk import BRSR data from Excel/CSV file. Maps columns to BRSR datapoints."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    # Get org
    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    # Read file
    content = await file.read()
    filename = file.filename or ""

    if filename.endswith(".csv"):
        import csv
        reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use .xlsx or .csv")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    # Map and import rows
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows):
        datapoint_id = row.get("datapoint_id") or row.get("Datapoint ID") or row.get("id")
        value = row.get("value") or row.get("Value") or row.get("Response")
        section = row.get("section") or row.get("Section") or "section_c"
        subsection = row.get("subsection") or row.get("Subsection") or "general"

        if not datapoint_id or value is None:
            skipped += 1
            continue

        # Normalize section
        section = section.lower().replace(" ", "_")
        if section not in ("section_a", "section_b", "section_c"):
            section = "section_c"

        try:
            entry_data = {
                "user_id": user_id,
                "org_id": org_id,
                "financial_year": financial_year,
                "section": section,
                "subsection": subsection,
                "datapoint_id": str(datapoint_id).strip(),
                "value": {"value": str(value).strip()} if not isinstance(value, dict) else value,
                "source": "imported",
                "notes": f"Bulk imported from {filename}",
            }

            sb.table("brsr_entries").upsert(
                entry_data,
                on_conflict="org_id,user_id,financial_year,datapoint_id"
            ).execute()
            imported += 1
        except Exception as e:
            errors.append({"row": i + 2, "datapoint_id": datapoint_id, "error": str(e)})

    return {
        "status": "complete",
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10],  # limit error list
        "total_rows": len(rows),
    }


@router.get("/import/template")
async def download_import_template(authorization: str = Header(...)):
    """Download an Excel template pre-filled with all BRSR datapoint IDs."""
    await get_user_id(authorization)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.brsr_datapoints import BRSR_DATAPOINTS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BRSR Data Import"

    # Headers
    headers = ["datapoint_id", "section", "subsection", "label", "value", "unit", "notes"]
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Fill datapoints
    row_idx = 2
    for dp in BRSR_DATAPOINTS:
        ws.cell(row=row_idx, column=1, value=dp.get("id", ""))
        ws.cell(row=row_idx, column=2, value=dp.get("section", "section_c"))
        ws.cell(row=row_idx, column=3, value=dp.get("subsection", ""))
        ws.cell(row=row_idx, column=4, value=dp.get("label", ""))
        ws.cell(row=row_idx, column=5, value="")  # user fills this
        ws.cell(row=row_idx, column=6, value=dp.get("unit", ""))
        ws.cell(row=row_idx, column=7, value="")
        row_idx += 1

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="FileBRSR_Import_Template.xlsx"'},
    )


# ═══════════════════════════════════════════════════════════════
# COMPARATIVE YoY ANALYSIS
# ═══════════════════════════════════════════════════════════════

@router.get("/analysis/yoy")
async def get_yoy_analysis(
    authorization: str = Header(...),
    financial_year: str = Query(default="FY2024-25"),
):
    """Compare current year vs previous year BRSR data. Returns trends and deltas."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    # Current year
    query = sb.table("brsr_entries").select("datapoint_id, value, section").eq("financial_year", financial_year)
    if org_id:
        query = query.eq("org_id", org_id)
    else:
        query = query.eq("user_id", user_id)
    current = query.execute()

    # Previous year
    prev_fy = _get_previous_fy(financial_year)
    prev_query = sb.table("brsr_entries").select("datapoint_id, value, section").eq("financial_year", prev_fy)
    if org_id:
        prev_query = prev_query.eq("org_id", org_id)
    else:
        prev_query = prev_query.eq("user_id", user_id)
    previous = prev_query.execute()

    # Build comparison
    current_map = {e["datapoint_id"]: e for e in (current.data or [])}
    previous_map = {e["datapoint_id"]: e for e in (previous.data or [])}

    all_datapoints = set(list(current_map.keys()) + list(previous_map.keys()))
    comparisons = []

    for dp_id in sorted(all_datapoints):
        curr_entry = current_map.get(dp_id)
        prev_entry = previous_map.get(dp_id)

        curr_val = _extract_numeric(curr_entry["value"]) if curr_entry else None
        prev_val = _extract_numeric(prev_entry["value"]) if prev_entry else None

        delta = None
        delta_pct = None
        trend = "unchanged"

        if curr_val is not None and prev_val is not None and prev_val != 0:
            delta = curr_val - prev_val
            delta_pct = round(((curr_val - prev_val) / abs(prev_val)) * 100, 2)
            trend = "improved" if delta_pct > 0 else "declined" if delta_pct < 0 else "unchanged"
        elif curr_val is not None and prev_val is None:
            trend = "new"
        elif curr_val is None and prev_val is not None:
            trend = "missing"

        comparisons.append({
            "datapoint_id": dp_id,
            "section": (curr_entry or prev_entry or {}).get("section", ""),
            "current_year": financial_year,
            "current_value": curr_val,
            "previous_year": prev_fy,
            "previous_value": prev_val,
            "delta": delta,
            "delta_pct": delta_pct,
            "trend": trend,
        })

    # Summary stats
    improved = len([c for c in comparisons if c["trend"] == "improved"])
    declined = len([c for c in comparisons if c["trend"] == "declined"])
    new_disclosures = len([c for c in comparisons if c["trend"] == "new"])
    missing = len([c for c in comparisons if c["trend"] == "missing"])

    return {
        "financial_year": financial_year,
        "previous_year": prev_fy,
        "summary": {
            "total_datapoints": len(comparisons),
            "improved": improved,
            "declined": declined,
            "unchanged": len([c for c in comparisons if c["trend"] == "unchanged"]),
            "new_disclosures": new_disclosures,
            "missing_this_year": missing,
        },
        "comparisons": comparisons,
    }


# ═══════════════════════════════════════════════════════════════
# BOARD-LEVEL EXECUTIVE DASHBOARD
# ═══════════════════════════════════════════════════════════════

@router.get("/board/dashboard")
async def get_board_dashboard(
    authorization: str = Header(...),
    financial_year: str = Query(default="FY2024-25"),
):
    """
    CXO-level executive summary. Single API call returns everything a board needs.
    No data-entry complexity — just KPIs, risks, and compliance status.
    """
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    # 1. Compliance Completion
    query = sb.table("brsr_entries").select("datapoint_id, section, value, source, verified").eq("financial_year", financial_year)
    if org_id:
        query = query.eq("org_id", org_id)
    else:
        query = query.eq("user_id", user_id)
    entries = query.execute()
    entry_data = entries.data or []

    total_mandatory = 216  # BRSR full
    filled = len(entry_data)
    verified = len([e for e in entry_data if e.get("verified")])
    ai_extracted = len([e for e in entry_data if e.get("source") == "ai_extracted"])

    # Section breakdown
    section_a = len([e for e in entry_data if e.get("section") == "section_a"])
    section_b = len([e for e in entry_data if e.get("section") == "section_b"])
    section_c = len([e for e in entry_data if e.get("section") == "section_c"])

    # 2. Risk Indicators
    # Suppliers at risk
    supplier_query = sb.table("suppliers").select("id, risk_level")
    if org_id:
        supplier_query = supplier_query.eq("org_id", org_id)
    else:
        supplier_query = supplier_query.eq("user_id", user_id)
    suppliers = supplier_query.execute()
    supplier_data = suppliers.data or []
    high_risk_suppliers = len([s for s in supplier_data if s.get("risk_level") in ("high", "critical")])

    # Compliance status
    compliance_query = sb.table("regulatory_compliance").select("regulation, status, due_date").eq("financial_year", financial_year)
    if org_id:
        compliance_query = compliance_query.eq("org_id", org_id)
    else:
        compliance_query = compliance_query.eq("user_id", user_id)
    compliances = compliance_query.execute()
    compliance_data = compliances.data or []
    non_compliant = len([c for c in compliance_data if c.get("status") == "non_compliant"])
    overdue = len([c for c in compliance_data if c.get("due_date") and c["due_date"] < str(date.today()) and c.get("status") not in ("compliant", "not_applicable")])

    # 3. Workflow Status
    workflow_query = sb.table("workflow_instances").select("status")
    if org_id:
        pass  # would filter by org, simplified for now
    workflow_instances = workflow_query.eq("status", "pending").execute()
    pending_approvals = len(workflow_instances.data or [])

    # 4. YoY Summary (quick version)
    prev_fy = _get_previous_fy(financial_year)
    prev_query = sb.table("brsr_entries").select("id").eq("financial_year", prev_fy)
    if org_id:
        prev_query = prev_query.eq("org_id", org_id)
    else:
        prev_query = prev_query.eq("user_id", user_id)
    prev_entries = prev_query.execute()
    prev_filled = len(prev_entries.data or [])
    disclosure_improvement = filled - prev_filled

    # 5. ESG Score Summary
    esg_query = sb.table("esg_ratings").select("agency, readiness_score").eq("financial_year", financial_year)
    if org_id:
        esg_query = esg_query.eq("org_id", org_id)
    else:
        esg_query = esg_query.eq("user_id", user_id)
    esg_ratings = esg_query.execute()

    # 6. Key deadlines coming up
    upcoming_deadlines = [
        c for c in compliance_data
        if c.get("due_date") and c["due_date"] >= str(date.today())
        and c.get("status") not in ("compliant", "not_applicable")
    ]
    upcoming_deadlines.sort(key=lambda x: x.get("due_date", ""))

    return {
        "financial_year": financial_year,
        "compliance_score": round((filled / total_mandatory) * 100, 1) if total_mandatory > 0 else 0,
        "completion": {
            "total_required": total_mandatory,
            "filled": filled,
            "verified": verified,
            "ai_extracted": ai_extracted,
            "manual": filled - ai_extracted,
            "completion_pct": round((filled / total_mandatory) * 100, 1),
            "verification_pct": round((verified / max(filled, 1)) * 100, 1),
        },
        "section_progress": {
            "section_a": {"filled": section_a, "total": 30, "pct": round((section_a / 30) * 100)},
            "section_b": {"filled": section_b, "total": 58, "pct": round((section_b / 58) * 100)},
            "section_c": {"filled": section_c, "total": 128, "pct": round((section_c / 128) * 100)},
        },
        "risks": {
            "high_risk_suppliers": high_risk_suppliers,
            "total_suppliers": len(supplier_data),
            "non_compliant_regulations": non_compliant,
            "overdue_filings": overdue,
            "pending_approvals": pending_approvals,
        },
        "yoy": {
            "previous_year": prev_fy,
            "prev_filled": prev_filled,
            "disclosure_improvement": disclosure_improvement,
            "improvement_pct": round((disclosure_improvement / max(prev_filled, 1)) * 100, 1) if prev_filled > 0 else None,
        },
        "esg_ratings": esg_ratings.data or [],
        "upcoming_deadlines": upcoming_deadlines[:5],
        "filing_readiness": {
            "ready": filled >= total_mandatory * 0.9 and verified >= filled * 0.8,
            "blockers": _get_filing_blockers(filled, verified, total_mandatory, non_compliant, pending_approvals),
        },
    }


# ═══════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def _get_previous_fy(financial_year: str) -> str:
    """Convert 'FY2024-25' to 'FY2023-24'."""
    try:
        # Handle formats like 'FY2024-25' or 'FY 2024-25'
        fy_clean = financial_year.replace(" ", "").replace("FY", "")
        parts = fy_clean.split("-")
        start = int(parts[0])
        return f"FY{start - 1}-{str(start)[-2:]}"
    except (ValueError, IndexError):
        return "FY2023-24"


def _extract_numeric(value) -> Optional[float]:
    """Extract a numeric value from various BRSR value formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        v = value.get("value")
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).replace(",", "").strip())
        except (ValueError, TypeError):
            return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _get_filing_blockers(filled: int, verified: int, total: int, non_compliant: int, pending: int) -> List[str]:
    """Identify what's blocking a filing submission."""
    blockers = []
    if filled < total * 0.9:
        blockers.append(f"Incomplete: {total - filled} datapoints still missing")
    if verified < filled * 0.5:
        blockers.append(f"Unverified: {filled - verified} entries need verification")
    if non_compliant > 0:
        blockers.append(f"{non_compliant} regulation(s) non-compliant")
    if pending > 0:
        blockers.append(f"{pending} workflow item(s) pending approval")
    return blockers


# ═══════════════════════════════════════════════════════════════
# LEAD CAPTURE (GTM Funnel)
# ═══════════════════════════════════════════════════════════════

class LeadCapture(BaseModel):
    email: str
    company_name: Optional[str] = None
    contact_name: Optional[str] = None
    source: str = "readiness_assessment"
    score: Optional[int] = None
    readiness_level: Optional[str] = None
    tags: Optional[List[str]] = None
    answers: Optional[dict] = None
    phase_scores: Optional[dict] = None
    metadata: Optional[dict] = None


@router.post("/leads/capture")
async def capture_lead(lead: LeadCapture):
    """Capture lead from readiness assessment, resource downloads, etc. No auth required."""
    sb = get_supabase_admin()

    data = {
        "email": lead.email,
        "company_name": lead.company_name,
        "contact_name": lead.contact_name,
        "source": lead.source,
        "score": lead.score,
        "readiness_level": lead.readiness_level,
        "tags": lead.tags,
        "metadata": lead.metadata or {
            "answers": lead.answers,
            "phase_scores": lead.phase_scores,
        },
        "created_at": datetime.utcnow().isoformat(),
    }

    try:
        result = sb.table("leads").insert(data).execute()
        lead_id = result.data[0]["id"] if result.data else None
    except Exception:
        # Table might not exist yet — that's fine, still return success
        lead_id = None

    # Send welcome/results email if we have Resend configured
    try:
        if lead.source == "readiness_assessment" and lead.score is not None:
            await send_email(
                to=lead.email,
                template_name="readiness_report",
                variables={
                    "name": lead.contact_name or "there",
                    "company": lead.company_name or "your company",
                    "score": lead.score,
                    "readiness_level": lead.readiness_level or "Unknown",
                    "phase_scores": lead.phase_scores or {},
                },
            )
        elif lead.source == "pilot_application":
            metadata = lead.metadata or {}
            await send_email(
                to=ADMIN_EMAIL,
                template_name="admin_pilot_notification",
                variables={
                    "company_name": lead.company_name or "Unknown",
                    "contact_name": lead.contact_name or "Unknown",
                    "email": lead.email,
                    "designation": metadata.get("designation", "N/A") if isinstance(metadata, dict) else "N/A",
                    "cin": metadata.get("cin", "N/A") if isinstance(metadata, dict) else "N/A",
                    "market_cap_range": metadata.get("market_cap_range", "N/A") if isinstance(metadata, dict) else "N/A",
                    "current_method": metadata.get("current_method", "N/A") if isinstance(metadata, dict) else "N/A",
                    "pain_points": metadata.get("pain_points", "N/A") if isinstance(metadata, dict) else "N/A",
                    "submitted_at": datetime.utcnow().strftime("%d %b %Y, %H:%M UTC"),
                },
            )
    except Exception:
        pass  # Non-blocking

    return {"status": "captured", "id": lead_id}


# ═══════════════════════════════════════════════════════════════
# SUPPLIER INVITE — THE VIRAL LOOP (Phase 2 Seed)
# ═══════════════════════════════════════════════════════════════

class SupplierInvite(BaseModel):
    supplier_name: str
    supplier_email: str
    contact_person: Optional[str] = None
    industry: Optional[str] = None
    tier: str = "tier_1"  # tier_1, tier_2, tier_3


class BulkSupplierInvite(BaseModel):
    suppliers: List[SupplierInvite]


@router.post("/suppliers/invite")
async def invite_supplier(invite: SupplierInvite, authorization: str = Header(...)):
    """Invite a single supplier for ESG assessment. Sends email with unique assessment link."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    # Get inviter's org
    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    # Get company name for the invite email
    profile = sb.table("profiles").select("company_name").eq("id", user_id).single().execute()
    buyer_company = profile.data.get("company_name", "A listed company") if profile.data else "A listed company"

    # Generate unique assessment token
    import secrets
    assessment_token = secrets.token_urlsafe(24)

    # Store supplier invite
    supplier_data = {
        "org_id": org_id,
        "invited_by": user_id,
        "supplier_name": invite.supplier_name,
        "supplier_email": invite.supplier_email,
        "contact_person": invite.contact_person,
        "industry": invite.industry,
        "tier": invite.tier,
        "assessment_token": assessment_token,
        "status": "invited",
        "invited_at": datetime.utcnow().isoformat(),
    }

    try:
        result = sb.table("supplier_invites").insert(supplier_data).execute()
        invite_id = result.data[0]["id"] if result.data else None
    except Exception:
        invite_id = None

    # Send assessment invite email
    assessment_url = f"https://filebrsr.com/assess/{assessment_token}"
    try:
        await send_email(
            to_email=invite.supplier_email,
            template="supplier_invite",
            variables={
                "supplier_name": invite.supplier_name,
                "buyer_company": buyer_company,
                "assessment_url": assessment_url,
                "contact_person": invite.contact_person or invite.supplier_name,
            },
        )
    except Exception:
        pass

    return {
        "status": "invited",
        "id": invite_id,
        "assessment_url": assessment_url,
        "token": assessment_token,
    }


@router.post("/suppliers/invite-bulk")
async def invite_suppliers_bulk(body: BulkSupplierInvite, authorization: str = Header(...)):
    """Invite multiple suppliers at once. CSV import on frontend sends here."""
    user_id = await get_user_id(authorization)
    results = []
    for supplier in body.suppliers[:100]:  # Cap at 100 per request
        try:
            result = await invite_supplier(supplier, authorization)
            results.append({"supplier": supplier.supplier_email, "status": "invited"})
        except Exception as e:
            results.append({"supplier": supplier.supplier_email, "status": "failed", "error": str(e)})

    return {
        "status": "bulk_complete",
        "total": len(body.suppliers),
        "invited": len([r for r in results if r["status"] == "invited"]),
        "failed": len([r for r in results if r["status"] == "failed"]),
        "results": results,
    }


@router.get("/suppliers/invited")
async def get_invited_suppliers(authorization: str = Header(...), status: Optional[str] = None):
    """Get all suppliers invited by this org with assessment status."""
    user_id = await get_user_id(authorization)
    sb = get_supabase_admin()

    org_member = sb.table("org_members").select("org_id").eq("user_id", user_id).maybe_single().execute()
    org_id = org_member.data.get("org_id") if org_member.data else None

    query = sb.table("supplier_invites").select("*").order("invited_at", desc=True)
    if org_id:
        query = query.eq("org_id", org_id)
    else:
        query = query.eq("invited_by", user_id)

    if status:
        query = query.eq("status", status)

    result = query.limit(200).execute()

    # Calculate stats
    suppliers = result.data or []
    stats = {
        "total_invited": len(suppliers),
        "completed": len([s for s in suppliers if s.get("status") == "completed"]),
        "pending": len([s for s in suppliers if s.get("status") == "invited"]),
        "in_progress": len([s for s in suppliers if s.get("status") == "in_progress"]),
    }

    return {"suppliers": suppliers, "stats": stats}
