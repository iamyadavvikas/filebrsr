"""
Generate comprehensive Excel workbook with all BRSR data points,
explanations, and ESRS cross-references.
"""
import sys
sys.path.insert(0, ".")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.brsr_datapoints import BRSR_DATAPOINTS, get_datapoints_stats, DATA_TYPES

# Explanations for each data point - how it satisfies BRSR compliance
EXPLANATIONS = {
    # Section A - General Disclosures
    "A.I.1": "The Corporate Identity Number uniquely identifies the listed entity under MCA. Required for regulatory identification and verification of BRSR filing entity.",
    "A.I.2": "Legal name of the entity as registered. Essential baseline identifier for all BRSR disclosures and cross-referencing with stock exchange filings.",
    "A.I.3": "Year of incorporation establishes the entity's operational history and maturity, relevant for assessing sustainability journey and trajectory.",
    "A.I.4": "Registered office address establishes legal domicile for regulatory correspondence and jurisdiction-specific compliance requirements.",
    "A.I.5": "Corporate/principal business address for stakeholder communication and geographical context of primary operations.",
    "A.I.6": "Official contact email for stakeholder queries related to sustainability/BRSR disclosures.",
    "A.I.7": "Official telephone contact for accessibility and stakeholder engagement on BRSR matters.",
    "A.I.8": "Company website where BRSR report and related policies are publicly accessible, ensuring transparency.",
    "A.I.9": "Identifies the specific financial year being reported, ensuring temporal accuracy and comparability of disclosures.",
    "A.I.10": "Stock exchange listing details establish the regulatory framework (SEBI LODR) under which BRSR is mandated.",
    "A.I.11": "Paid-up capital indicates the entity's scale and establishes size-based applicability of certain BRSR requirements.",
    "A.I.12": "Designated contact person ensures accountability and provides stakeholders a channel for BRSR-related queries.",
    "A.I.13": "Reporting boundary (standalone/consolidated) defines the scope of all subsequent disclosures. Critical for BRSR Core assurance.",
    "A.I.14": "Assurance provider details establish credibility of BRSR Core disclosures. Required when entity opts for/mandated assurance.",
    "A.I.15": "Type of assurance (limited/reasonable) indicates the level of confidence stakeholders can place in verified data points.",
    "A.II.1": "Business activity details (top 90% of turnover) establishes the entity's core operations and sector classification for materiality assessment.",
    "A.II.2": "Products/services with NIC codes enable sector-specific analysis and comparability across companies in same industry.",
    "A.II.3": "NIC codes provide standardized industrial classification for benchmarking and sector-specific BRSR expectations.",
    "A.III.1": "National plant locations indicate geographical spread of manufacturing/processing operations and related environmental/social impacts.",
    "A.III.2": "International plant locations indicate global operational footprint and cross-border sustainability considerations.",
    "A.III.3": "National office locations indicate administrative presence across states for employment and community impact assessment.",
    "A.III.4": "International office locations indicate global services/administrative presence.",
    "A.III.5": "Number of states/UTs served shows domestic market coverage and geographic diversity of stakeholder impact.",
    "A.III.6": "Number of countries served shows international reach and applicability of global sustainability standards.",
    "A.III.7": "Export percentage indicates exposure to international sustainability requirements and global supply chain standards.",
    "A.III.8": "Customer types (B2B/B2C/Government) affect consumer responsibility disclosures under Principle 9.",
    "A.IV.1": "Male permanent employees - workforce composition data essential for gender diversity and equal opportunity analysis under BRSR Core.",
    "A.IV.2": "Female permanent employees - critical for assessing gender balance in permanent workforce (BRSR Core indicator).",
    "A.IV.3": "Total permanent employees establishes baseline for all per-employee calculations (training %, benefits coverage).",
    "A.IV.4": "Male contractual/temporary workforce - captures non-permanent workforce often facing different working conditions.",
    "A.IV.5": "Female contractual/temporary workforce - gender breakdown of non-permanent staff for equity assessment.",
    "A.IV.6": "Total non-permanent employees for full workforce picture including contract/temporary/seasonal staff.",
    "A.IV.7": "Male permanent workers (non-managerial) - BRSR Core requires gender-disaggregated worker data separate from employees.",
    "A.IV.8": "Female permanent workers - critical for gender equity assessment in operational/blue-collar workforce.",
    "A.IV.9": "Total permanent workers establishes baseline for worker-specific metrics (safety incidents, training).",
    "A.IV.10": "Male non-permanent workers - contract labour often at higher safety/rights risk, requires separate tracking.",
    "A.IV.11": "Female non-permanent workers - intersectional vulnerability (gender + employment type) monitoring.",
    "A.IV.12": "Total non-permanent workers for complete operational workforce coverage.",
    "A.IV.13": "Differently abled permanent employees data ensures compliance with disability inclusion requirements and Rights of Persons with Disabilities Act.",
    "A.IV.14": "Differently abled non-permanent employees tracks inclusivity across employment types.",
    "A.IV.15": "Differently abled permanent workers tracks inclusion at operational level.",
    "A.IV.16": "Differently abled non-permanent workers ensures disability inclusion across all workforce categories.",
    "A.IV.17": "Women on Board (%) - BRSR Core indicator measuring gender diversity at highest governance level. SEBI mandates minimum one woman director.",
    "A.IV.18": "Women in KMP (%) measures gender representation in executive management decision-making.",
    "A.IV.19": "Male employee turnover rate (3-year trend) indicates workforce stability, satisfaction, and retention effectiveness.",
    "A.IV.20": "Female employee turnover rate (3-year trend) helps identify gender-specific retention challenges.",
    "A.IV.21": "Male worker turnover rate indicates operational workforce stability and working conditions quality.",
    "A.IV.22": "Female worker turnover rate helps identify if women workers face disproportionate attrition.",
    "A.V.1": "Group company details establish the corporate structure and scope of sustainability governance across entities.",
    "A.V.2": "Extension of ethics policies to group companies demonstrates responsible governance across corporate structure.",
    "A.VI.1": "CSR applicability determines whether Section 135 spending disclosures are required under Principle 8.",
    "A.VI.2": "Turnover establishes financial scale for intensity calculations (energy/water/waste per rupee of turnover).",
    "A.VI.3": "Net worth used alongside turnover for CSR threshold calculation and financial context.",
    "A.VII.1": "Complaints/grievances on each NGRBC principle measures stakeholder feedback and redressal effectiveness.",
    "A.VII.2": "Material ESG issues overview provides context for subsequent principle-wise disclosures and materiality linkage.",

    # Section B - Management and Process
    "B.1": "Policy coverage across all 9 NGRBC principles demonstrates systematic governance approach to responsible business conduct.",
    "B.2": "Board approval of policies establishes highest-level accountability and governance commitment to sustainability.",
    "B.3": "Public web links to policies ensure transparency and stakeholder access to entity's sustainability commitments.",
    "B.4": "Translation of policies into procedures ensures practical implementation, not just documentation of intent.",
    "B.5": "Value chain extension of policies demonstrates responsible supply chain management beyond entity's own operations.",
    "B.6": "Certifications/standards (ISO 14001, SA 8000, etc.) provide third-party validation of sustainability management systems.",
    "B.7": "Specific goals with timelines demonstrate measurable commitment and enable progress tracking by stakeholders.",
    "B.8": "Performance against targets provides accountability and demonstrates actual progress on sustainability commitments.",
    "B.9": "Director's statement on ESG establishes tone from the top and strategic priority of sustainability.",
    "B.10": "Designated authority for BR implementation ensures clear accountability for sustainability performance.",
    "B.11": "Board Committee consideration of sustainability ensures governance-level oversight of ESG risks and opportunities.",
    "B.12": "Regular NGRBC review by Board/Committee demonstrates ongoing governance oversight, not one-time compliance.",
    "B.13": "External assessment of policies provides independent validation of sustainability governance effectiveness.",
    "B.14": "Name of external assessor ensures transparency about independence and credibility of policy evaluation.",

    # Section C - Principle 1
    "C.P1.E.1": "Board training on NGRBC principles ensures directors understand their sustainability governance responsibilities.",
    "C.P1.E.2": "KMP training ensures executive leadership can embed sustainability in business decisions and operations.",
    "C.P1.E.3": "Employee training creates organizational awareness and embeds ethical conduct across the workforce.",
    "C.P1.E.4": "Worker training extends ethical awareness to operational level where many compliance risks materialize.",
    "C.P1.E.5": "Fines/penalties disclosure - BRSR Core indicator revealing regulatory non-compliance that signals governance failures.",
    "C.P1.E.6": "Appeal details show entity's response to regulatory action and whether issues are contested or accepted.",
    "C.P1.E.7": "Anti-corruption/bribery policy existence - BRSR Core indicator demonstrating zero-tolerance stance on corruption.",
    "C.P1.E.8": "Disciplinary actions for bribery/corruption demonstrate that anti-corruption policies have enforcement mechanisms.",
    "C.P1.E.9": "Director conflict of interest complaints track governance integrity at highest level.",
    "C.P1.E.10": "KMP conflict of interest complaints track executive-level integrity and independence.",
    "C.P1.E.11": "Corrective actions on penalties show entity's responsiveness and commitment to preventing recurrence.",
    "C.P1.E.12": "Accounts payable days indicate fair treatment of suppliers and ethical business practices (not delaying payments).",
    "C.P1.E.13": "Purchase concentration reveals supply chain dependency risks and potential for unfair trade influence.",
    "C.P1.E.14": "Sales concentration reveals customer dependency and market power dynamics relevant to fair business practices.",
    "C.P1.L.1": "Value chain awareness extends ethical standards beyond entity boundary, promoting responsible business ecosystem.",
    "C.P1.L.2": "Conflict of interest avoidance processes show proactive governance rather than reactive complaint handling.",

    # Section C - Principle 2
    "C.P2.E.1": "Environmental R&D/capex (% of turnover) shows investment commitment to green product innovation.",
    "C.P2.E.2": "Social R&D/capex shows investment in product improvements benefiting society (accessibility, safety, etc.).",
    "C.P2.E.3": "Sustainable sourcing procedures - BRSR Core indicator for responsible procurement and supply chain stewardship.",
    "C.P2.E.4": "% inputs sourced sustainably quantifies progress on sustainable procurement commitments.",
    "C.P2.E.5": "End-of-life product reclaim processes demonstrate circular economy commitment and extended producer responsibility.",
    "C.P2.E.6": "EPR applicability determines whether entity must meet specific plastic/e-waste collection/recycling targets.",
    "C.P2.E.7": "EPR plastic waste data demonstrates compliance with Plastic Waste Management Rules collection/recycling targets.",
    "C.P2.E.8": "EPR e-waste data demonstrates compliance with E-Waste Management Rules collection/recycling targets.",
    "C.P2.L.1": "LCA adoption shows advanced lifecycle thinking for products — identifying full environmental/social footprint.",
    "C.P2.L.2": "LCA results identify hotspots for improvement and demonstrate evidence-based sustainability strategy.",
    "C.P2.L.3": "Recycled input material % - BRSR Core leadership indicator measuring circularity in production.",
    "C.P2.L.4": "Reuse of products/packaging demonstrates highest-value circular economy practice (above recycling).",
    "C.P2.L.5": "Recycling of products/packaging demonstrates resource recovery at end of product life.",

    # Section C - Principle 3
    "C.P3.E.1": "Health insurance coverage (%) for employees - BRSR Core well-being indicator for social protection.",
    "C.P3.E.2": "Accident insurance coverage ensures financial protection for workers facing occupational hazards.",
    "C.P3.E.3": "Maternity benefits coverage (%) ensures compliance with Maternity Benefit Act and women's workplace rights.",
    "C.P3.E.4": "Paternity benefits coverage supports gender-equal caregiving and progressive workplace policies.",
    "C.P3.E.5": "Day care facility coverage supports working parents and women's workforce participation.",
    "C.P3.E.6": "Worker well-being measures ensure non-managerial/operational workforce receives equivalent social protections.",
    "C.P3.E.7": "Well-being spending as % of revenue quantifies entity's investment in human capital beyond statutory minimums.",
    "C.P3.E.8": "Retirement benefits (PF/Gratuity/ESI) timely deposit ensures compliance with social security laws and worker financial security.",
    "C.P3.E.9": "Workplace accessibility for differently abled ensures compliance with Rights of Persons with Disabilities Act.",
    "C.P3.E.10": "Equal opportunity policy demonstrates commitment to non-discrimination in employment.",
    "C.P3.E.11": "Parental leave return/retention rates measure whether family-friendly policies translate to actual workforce retention.",
    "C.P3.E.12": "Employee grievance mechanism ensures workers have accessible channels to raise workplace concerns without retaliation.",
    "C.P3.E.13": "Worker grievance mechanism extends redressal access to operational/contractual workforce.",
    "C.P3.E.14": "Union membership data indicates freedom of association and collective bargaining rights recognition.",
    "C.P3.E.15": "Health & safety training (by gender) - BRSR Core indicator ensuring all workers are equipped to maintain safe workplaces.",
    "C.P3.E.16": "Skill upgradation training shows investment in workforce development and future employability.",
    "C.P3.E.17": "Employee performance reviews measure access to career development and transparent evaluation.",
    "C.P3.E.18": "Worker performance reviews extend career development access beyond white-collar employees.",
    "C.P3.E.19": "OHS management system - BRSR Core indicator confirming systematic approach to workplace safety (ISO 45001 or equivalent).",
    "C.P3.E.20": "LTIFR (Lost Time Injury Frequency Rate) - BRSR Core safety performance metric for benchmarking.",
    "C.P3.E.21": "Fatalities - BRSR Core indicator. Any workplace death signals critical safety failures requiring disclosure.",
    "C.P3.E.22": "Reportable injuries track serious safety incidents requiring regulatory notification.",
    "C.P3.E.23": "Safe workplace measures describe specific actions taken — engineering controls, PPE, training, monitoring.",
    "C.P3.E.24": "Working condition complaints - BRSR Core indicator tracking worker-raised safety/health concerns and resolution.",
    "C.P3.E.25": "% plants/offices assessed for H&S shows extent of systematic safety auditing across operations.",
    "C.P3.L.1": "Life insurance/disability cover beyond statutory minimum shows enhanced worker protection.",
    "C.P3.L.2": "Transition assistance for retiring workers demonstrates responsible employer practices.",
    "C.P3.L.3": "High consequence injuries track severe injuries requiring extended recovery — beyond LTIFR.",
    "C.P3.L.4": "Rehabilitation support for injured workers shows duty of care beyond immediate medical treatment.",

    # Section C - Principle 4
    "C.P4.E.1": "Stakeholder identification demonstrates structured approach to understanding who is affected by operations.",
    "C.P4.E.2": "Engagement channels/platforms show how entity maintains ongoing dialogue with diverse stakeholder groups.",
    "C.P4.L.1": "Formal consultation processes show stakeholder voice influences actual business decisions.",
    "C.P4.L.2": "Stakeholder concerns in decision-making demonstrates genuine responsiveness, not just information sharing.",

    # Section C - Principle 5
    "C.P5.E.1": "HR training for employees (%) - BRSR Core indicator ensuring workforce understands human rights obligations.",
    "C.P5.E.2": "HR training for workers extends human rights awareness to operational workforce often at higher vulnerability.",
    "C.P5.E.3": "Minimum wage compliance for employees - BRSR Core indicator for living wage and labour law compliance.",
    "C.P5.E.4": "Minimum wage compliance for workers - often contract/gig workers face wage violations, this tracks compliance.",
    "C.P5.E.5": "Median remuneration gender comparison - BRSR Core indicator measuring gender pay equity across levels.",
    "C.P5.E.6": "Female wages as % of total wages quantifies gender pay distribution at each organizational level.",
    "C.P5.E.7": "Sexual harassment complaints (POSH) track compliance with Prevention of Sexual Harassment Act.",
    "C.P5.E.8": "Discrimination complaints track workplace equality and identify systemic bias issues.",
    "C.P5.E.9": "Child labour complaints - BRSR Core indicator. Zero tolerance required; any complaints signal severe HR failure.",
    "C.P5.E.10": "Forced/involuntary labour complaints - BRSR Core indicator for fundamental labour rights compliance.",
    "C.P5.E.11": "Wage complaints track disputes on payment, overtime, deductions — basic labour rights compliance.",
    "C.P5.E.12": "% operations assessed for HR shows systematic due diligence coverage across all locations.",
    "C.P5.E.13": "Corrective actions on HR risks demonstrate responsive action, not just identification of issues.",
    "C.P5.L.1": "Process modifications from HR assessments show organizational learning and systemic improvement.",
    "C.P5.L.2": "Value chain HR due diligence extends human rights responsibility to suppliers/contractors/subcontractors.",

    # Section C - Principle 6
    "C.P6.E.1": "Renewable energy consumption (GJ) - BRSR Core indicator measuring clean energy transition progress.",
    "C.P6.E.2": "Non-renewable energy consumption establishes baseline for decarbonization targets and fossil fuel dependency.",
    "C.P6.E.3": "Total electricity consumption enables granular analysis of energy mix and electrification opportunities.",
    "C.P6.E.4": "Total fuel consumption tracks direct combustion sources (diesel, gas, coal) for emission reduction planning.",
    "C.P6.E.5": "Energy intensity per turnover - BRSR Core efficiency metric enabling cross-company and year-on-year comparison.",
    "C.P6.E.6": "Renewable energy target disclosure - BRSR Core indicator showing voluntary climate commitment beyond compliance.",
    "C.P6.E.7": "PAT scheme participation indicates large energy consumer status and Bureau of Energy Efficiency compliance.",
    "C.P6.E.8": "Surface water withdrawal tracks dependency on rivers/lakes — sensitive to seasonal/climate water stress.",
    "C.P6.E.9": "Groundwater withdrawal - critical indicator given India's groundwater depletion crisis.",
    "C.P6.E.10": "Third-party water (municipal/tanker) indicates infrastructure dependency and indirect water source.",
    "C.P6.E.11": "Total water withdrawal - BRSR Core aggregate water footprint indicator.",
    "C.P6.E.12": "Total water consumption (withdrawal minus discharge) shows actual water consumption impact.",
    "C.P6.E.13": "Water intensity per turnover - BRSR Core efficiency indicator for water use productivity.",
    "C.P6.E.14": "Water discharge by destination/treatment ensures compliance with Water (Prevention & Control of Pollution) Act.",
    "C.P6.E.15": "Water stress area disclosure - BRSR Core indicator for climate adaptation and water risk assessment.",
    "C.P6.E.16": "Scope 1 GHG emissions - BRSR Core indicator. Direct emissions from owned/controlled sources.",
    "C.P6.E.17": "Scope 2 GHG emissions - BRSR Core indicator. Indirect emissions from purchased electricity/heat/steam.",
    "C.P6.E.18": "GHG intensity per turnover - BRSR Core metric for carbon efficiency and decarbonization progress.",
    "C.P6.E.19": "Voluntary GHG reduction target - BRSR Core indicator showing climate commitment (e.g., SBTi, Net Zero).",
    "C.P6.E.20": "CDM/voluntary mechanism project registration shows carbon market participation and verified reduction.",
    "C.P6.E.21": "Plastic waste generated - regulated under Plastic Waste Management Rules, EPR obligations.",
    "C.P6.E.22": "E-waste generated - regulated under E-Waste Management Rules, channelization obligations.",
    "C.P6.E.23": "Bio-medical waste - regulated under BMW Rules, requires authorized treatment/disposal.",
    "C.P6.E.24": "Construction/demolition waste - regulated under C&D Waste Management Rules.",
    "C.P6.E.25": "Battery waste - regulated under Battery Waste Management Rules 2022.",
    "C.P6.E.26": "Radioactive waste - regulated by AERB, requires specialized handling and disposal.",
    "C.P6.E.27": "Other hazardous waste - regulated under Hazardous Waste Management Rules, requires SPCB authorization.",
    "C.P6.E.28": "Non-hazardous waste covers general solid waste requiring responsible management and diversion from landfill.",
    "C.P6.E.29": "Total waste generated - BRSR Core aggregate indicator for overall waste footprint.",
    "C.P6.E.30": "Waste intensity per turnover - BRSR Core efficiency metric for waste reduction progress.",
    "C.P6.E.31": "Waste recycled tracks material recovery and circular economy practices.",
    "C.P6.E.32": "Waste reused tracks highest-value waste diversion (direct reuse without reprocessing).",
    "C.P6.E.33": "Other waste recovery (composting, co-processing) tracks alternative diversion methods.",
    "C.P6.E.34": "Waste to landfill - lowest-value disposal; reduction target indicator for circular economy.",
    "C.P6.E.35": "Waste incineration - tracks thermal treatment; relevant for emissions and energy recovery.",
    "C.P6.E.36": "Operations near ecologically sensitive areas requires heightened environmental management and biodiversity considerations.",
    "C.P6.E.37": "EIA details demonstrate compliance with EIA Notification 2006 and environmental clearance conditions.",
    "C.P6.E.38": "Environmental non-compliances track regulatory violations under Air/Water/Environment Acts and SPCB orders.",
    "C.P6.L.1": "Zero Liquid Discharge (ZLD) - BRSR Core leadership indicator for water stewardship excellence.",
    "C.P6.L.2": "Scope 3 GHG emissions - BRSR Core leadership indicator capturing full value chain carbon footprint.",
    "C.P6.L.3": "Voluntary environmental pledges show beyond-compliance commitment (e.g., EP100, RE100, SBTi).",
    "C.P6.L.4": "Biodiversity impact assessment required near protected areas; demonstrates ecological responsibility.",
    "C.P6.L.5": "Circular economy framework indicates systematic approach to resource efficiency beyond waste management.",

    # Section C - Principle 7
    "C.P7.E.1": "Industry association memberships reveal potential influence channels and policy engagement platforms.",
    "C.P7.E.2": "Anti-competitive conduct cases (5 years) reveal market behavior ethics and competition law compliance.",
    "C.P7.L.1": "Public policy positions show entity's advocacy stance on ESG issues (climate, labour, taxation).",

    # Section C - Principle 8
    "C.P8.E.1": "SIA details demonstrate due diligence for project impacts on communities (land acquisition, displacement).",
    "C.P8.E.2": "R&R (Rehabilitation & Resettlement) details show compliance with land acquisition laws and community care.",
    "C.P8.E.3": "Community grievance mechanism ensures affected communities have accessible voice for raising concerns.",
    "C.P8.E.4": "Community complaints data tracks volume and resolution of community-raised issues.",
    "C.P8.E.5": "Inputs from MSMEs/small producers (%) - BRSR Core indicator for inclusive supply chain and local economic impact.",
    "C.P8.E.6": "Local sourcing (%) - BRSR Core indicator measuring supply chain contribution to local/regional economy.",
    "C.P8.E.7": "CSR project details ensure compliance with Section 135/Schedule VII and demonstrate social investment impact.",
    "C.P8.L.1": "Job creation in smaller towns shows inclusive growth beyond metro/urban centres.",
    "C.P8.L.2": "Impact on aspirational districts aligns with government's inclusive development priorities.",
    "C.P8.L.3": "Preferential procurement for marginalized groups demonstrates proactive inclusion in supply chain.",
    "C.P8.L.4": "CSR beneficiary demographics (SC/ST/Others) track equity in social investment distribution.",

    # Section C - Principle 9
    "C.P9.E.1": "Data privacy complaints - BRSR Core indicator tracking consumer digital rights violations.",
    "C.P9.E.2": "Advertising complaints track misleading/unfair advertising practices affecting consumer choice.",
    "C.P9.E.3": "Cyber-security complaints - BRSR Core indicator for digital trust and data protection incidents.",
    "C.P9.E.4": "Essential service delivery complaints track consumer access and quality of essential services.",
    "C.P9.E.5": "Restrictive trade practice complaints reveal market behavior affecting consumer choice/pricing.",
    "C.P9.E.6": "Unfair trade practice complaints track deceptive/exploitative consumer interactions.",
    "C.P9.E.7": "Product recalls - BRSR Core indicator for product safety failures requiring market withdrawal.",
    "C.P9.E.8": "Recall reasons provide root cause transparency for product safety incidents.",
    "C.P9.E.9": "Cybersecurity/privacy policy existence confirms governance framework for consumer data protection.",
    "C.P9.E.10": "Corrective actions on consumer issues demonstrate responsive quality/service improvement.",
    "C.P9.L.1": "Environmental/social product information empowers consumer sustainable choice (eco-labels, carbon footprint).",
    "C.P9.L.2": "Consumer surveys/feedback mechanisms show proactive listening and service improvement commitment.",
    "C.P9.L.3": "Turnover from responsible products quantifies commercial integration of sustainability in product portfolio.",
}


def create_excel():
    wb = Workbook()
    
    # ==================== Sheet 1: All Data Points ====================
    ws = wb.active
    ws.title = "BRSR Data Points"
    
    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
    core_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    mandatory_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    voluntary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Headers
    headers = ["#", "Data Point ID", "Label", "Data Type", "Section", "Subsection",
               "Indicator Type", "Mandatory", "BRSR Core", "Conditional", 
               "ESRS Cross-Reference", "BRSR Paragraph", "Explanation / Compliance Rationale"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    for idx, dp in enumerate(BRSR_DATAPOINTS, 1):
        row = idx + 1
        explanation = EXPLANATIONS.get(dp["id"], "Disclosure required under SEBI BRSR framework for comprehensive sustainability reporting.")
        
        values = [
            idx,
            dp["id"],
            dp["label"],
            dp["data_type"].upper(),
            dp["section"].replace("section_", "Section ").upper(),
            dp["subsection"].replace("_", " ").title(),
            dp["indicator_type"].capitalize(),
            "Yes" if dp["mandatory"] else "No",
            "Yes" if dp["core"] else "No",
            "Yes" if dp["conditional"] else "No",
            dp["esrs_ref"] or "—",
            dp["paragraph_ref"],
            explanation,
        ]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 13))
            
            # Row coloring
            if dp["core"]:
                cell.fill = core_fill
            elif dp["mandatory"]:
                cell.fill = mandatory_fill
            elif not dp["mandatory"]:
                cell.fill = voluntary_fill
    
    # Column widths
    col_widths = [5, 12, 55, 14, 14, 22, 14, 12, 12, 12, 22, 18, 80]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # ==================== Sheet 2: Summary Statistics ====================
    ws2 = wb.create_sheet("Statistics")
    stats = get_datapoints_stats()
    
    # Title
    ws2.cell(row=1, column=1, value="BRSR Data Points — Summary Statistics").font = Font(size=14, bold=True)
    ws2.cell(row=2, column=1, value="Based on SEBI BRSR Annexure II, cross-referenced with EFRAG IG 3 ESRS").font = Font(size=10, italic=True)
    
    # Overview table
    ws2.cell(row=4, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=4, column=2, value="Count").font = Font(bold=True)
    
    overview = [
        ("Total Data Points", stats["total_datapoints"]),
        ("Mandatory (Shall)", stats["mandatory"]),
        ("Voluntary (Leadership)", stats["voluntary"]),
        ("BRSR Core (Assurance)", stats["core_assurance"]),
        ("ESRS Cross-referenced", stats["esrs_mapped"]),
        ("Conditional (If applicable)", stats["conditional"]),
    ]
    for i, (metric, count) in enumerate(overview, 5):
        ws2.cell(row=i, column=1, value=metric)
        ws2.cell(row=i, column=2, value=count)
    
    # By data type
    ws2.cell(row=13, column=1, value="Data Points by Type").font = Font(size=12, bold=True)
    ws2.cell(row=14, column=1, value="Data Type").font = Font(bold=True)
    ws2.cell(row=14, column=2, value="Count").font = Font(bold=True)
    ws2.cell(row=14, column=3, value="Description").font = Font(bold=True)
    
    for i, (dtype, count) in enumerate(sorted(stats["by_data_type"].items(), key=lambda x: -x[1]), 15):
        ws2.cell(row=i, column=1, value=dtype.capitalize())
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=DATA_TYPES.get(dtype, ""))
    
    # By section
    row_offset = 15 + len(stats["by_data_type"]) + 2
    ws2.cell(row=row_offset, column=1, value="Data Points by Section").font = Font(size=12, bold=True)
    section_names = {"section_a": "Section A: General Disclosures", "section_b": "Section B: Management & Process", "section_c": "Section C: Principle-wise Performance"}
    for i, (section, count) in enumerate(stats["by_section"].items(), row_offset + 1):
        ws2.cell(row=i, column=1, value=section_names.get(section, section))
        ws2.cell(row=i, column=2, value=count)
    
    # By principle
    row_offset += len(stats["by_section"]) + 3
    ws2.cell(row=row_offset, column=1, value="Data Points by Principle (Section C)").font = Font(size=12, bold=True)
    principle_names = {
        "principle_1": "P1: Ethics, Transparency & Accountability",
        "principle_2": "P2: Sustainable Products/Services",
        "principle_3": "P3: Employee Well-being",
        "principle_4": "P4: Stakeholder Engagement",
        "principle_5": "P5: Human Rights",
        "principle_6": "P6: Environment",
        "principle_7": "P7: Policy Advocacy",
        "principle_8": "P8: Inclusive Growth",
        "principle_9": "P9: Consumer Responsibility",
    }
    for i, (principle, count) in enumerate(stats["by_principle"].items(), row_offset + 1):
        ws2.cell(row=i, column=1, value=principle_names.get(principle, principle))
        ws2.cell(row=i, column=2, value=count)
    
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 55
    
    # ==================== Sheet 3: ESRS Mapping ====================
    ws3 = wb.create_sheet("ESRS Mapping")
    
    ws3.cell(row=1, column=1, value="BRSR ↔ ESRS Cross-Reference Mapping").font = Font(size=14, bold=True)
    ws3.cell(row=2, column=1, value="Mapping SEBI BRSR data points to European Sustainability Reporting Standards (ESRS)").font = Font(size=10, italic=True)
    
    esrs_headers = ["BRSR ID", "BRSR Data Point", "ESRS Reference", "ESRS Standard", "Mandatory", "BRSR Core"]
    for col, h in enumerate(esrs_headers, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row = 5
    for dp in BRSR_DATAPOINTS:
        if dp["esrs_ref"]:
            ws3.cell(row=row, column=1, value=dp["id"])
            ws3.cell(row=row, column=2, value=dp["label"])
            ws3.cell(row=row, column=3, value=dp["esrs_ref"])
            # Extract ESRS standard name
            esrs_std = dp["esrs_ref"].split(" ")[1].split("-")[0] if " " in dp["esrs_ref"] else dp["esrs_ref"]
            ws3.cell(row=row, column=4, value=esrs_std)
            ws3.cell(row=row, column=5, value="Yes" if dp["mandatory"] else "No")
            ws3.cell(row=row, column=6, value="Yes" if dp["core"] else "No")
            row += 1
    
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 12
    ws3.freeze_panes = "A5"

    # ==================== Sheet 4: Legend ====================
    ws4 = wb.create_sheet("Legend")
    ws4.cell(row=1, column=1, value="Color Legend & Definitions").font = Font(size=14, bold=True)
    
    ws4.cell(row=3, column=1, value="Row Colors in 'BRSR Data Points' sheet:").font = Font(bold=True)
    ws4.cell(row=4, column=1, value="Green background")
    ws4.cell(row=4, column=1).fill = core_fill
    ws4.cell(row=4, column=2, value="BRSR Core data point — subject to mandatory assurance")
    ws4.cell(row=5, column=1, value="Yellow background")
    ws4.cell(row=5, column=1).fill = mandatory_fill
    ws4.cell(row=5, column=2, value="Mandatory data point — required but not part of Core assurance")
    ws4.cell(row=6, column=1, value="Gray background")
    ws4.cell(row=6, column=1).fill = voluntary_fill
    ws4.cell(row=6, column=2, value="Voluntary/Leadership data point — disclose if applicable")
    
    ws4.cell(row=8, column=1, value="Definitions:").font = Font(bold=True)
    definitions = [
        ("Mandatory", "Must be disclosed by all listed entities filing BRSR"),
        ("BRSR Core", "Subject to reasonable/limited assurance as per SEBI circular. Includes specific KPIs across all principles."),
        ("Leadership Indicator", "Voluntary disclosure that demonstrates leadership in responsible business conduct"),
        ("Conditional", "Only required if a specific condition is met (e.g., EPR applicable, CSR applicable)"),
        ("ESRS Cross-reference", "Corresponding disclosure requirement in European Sustainability Reporting Standards"),
        ("Essential Indicator", "Mandatory disclosure under the relevant NGRBC principle"),
        ("Data Type", "Classification of the expected disclosure format (narrative text, numeric value, table, boolean, etc.)"),
    ]
    for i, (term, defn) in enumerate(definitions, 9):
        ws4.cell(row=i, column=1, value=term).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=defn)
    
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 90

    # Save
    output_path = "/Users/vikasyadav/Documents/Saas/filebrsr/BRSR_DataPoints_Complete.xlsx"
    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")
    print(f"  Sheet 1: BRSR Data Points ({len(BRSR_DATAPOINTS)} rows)")
    print(f"  Sheet 2: Statistics (summary counts)")
    print(f"  Sheet 3: ESRS Mapping ({sum(1 for dp in BRSR_DATAPOINTS if dp['esrs_ref'])} mapped)")
    print(f"  Sheet 4: Legend & Definitions")


if __name__ == "__main__":
    create_excel()
